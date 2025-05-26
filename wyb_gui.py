import time
import shutil
from sys_test import rotate_zp
from do_image import convert_new
import serial
import pymysql
import cv2
import numpy as np
import sys
import robot_act_new
import threading
import copy
import new_bzy
from do_image import testGrabImage
from do_image import ewm_and_lcd
from do_image import ewm_and_lcd_formal
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from PySide2.QtWidgets import QApplication, QTableWidget, QTableWidgetItem, QInputDialog, QMainWindow, QMessageBox, \
    QGraphicsScene, QFileDialog, QInputDialog, QPushButton, QLineEdit, QLabel, QPlainTextEdit, QDialog
from PySide2.QtUiTools import QUiLoader
from PySide2.QtGui import QPalette,QPixmap
from PySide2.QtWidgets import QApplication, QWidget, QPushButton, QVBoxLayout, QFileDialog
from PySide2.QtGui import QIcon
from PySide2 import QtCore
from PySide2.QtCore import QThread, Signal
from datetime import datetime
import os
import signal
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import pandas as pd
from openpyxl import Workbook

# 登录的路径
addr_ui = "C:/banshouGUI/ui/"

hang = 2

vector_wyb = []

wyb_num_all = ['FlUKE1','VICTORY1','KLET1','PLAY1','NEW1','TREATR1','ZHUYI1',]

wyb_style_all = ['FlUKE','VICTORY','KLET','PLAY','NEW','TREATR','ZHUYI',]

num_all = 2

equip_image = 'E:/photo/zp.jpg'

db = robot_act_new.db


vector_v1 = [1,2,3,4,5]
bzy_v1 = [1,2,3,4,5]
vector_v2 = [1,2,3,4,5]
bzy_v2 = [1,2,3,4,5]

vector_a1 = [1,2,3,4,5]
bzy_a1 = [1,2,3,4,5]
vector_a2 = [1,2,3,4,5]
bzy_a2 = [1,2,3,4,5]
vector_a3 = [1,2,3,4,5]
bzy_a3 = [1,2,3,4,5]
vector_a4 = [1,2,3,4,5]
bzy_a4 = [1,2,3,4,5]


find_wyb_way = '0'
find_result_name = '0'

be = 0
en = 1
k_num = 2

tool_x = -98.024
tool_y = -10.752

#be = 8
#en = 9
#k_num = 10

key_ok = 0
current_wyb = 0

buchang_x = -5
buchang_y = 12.7


photo_num = 1




class system_wyb():


    def __init__(self):

        self.ui = QUiLoader().load(addr_ui + "/system.ui")

        self.ui.system_wyb.clicked.connect(self.system_wyb)  # 下一步

        self.ui.system_exit.clicked.connect(self.system_exit)  # 下一步

        self.ui.system_test.clicked.connect(self.system_test)  # 下一步

        linshi_addr = "QLabel{border-image: url(" + addr_ui + "/登录底板.png)}"
        self.ui.system_pic.setStyleSheet(linshi_addr)

        self.ui.system_wyb.setVisible(False)

        self.ui.system_digital.setVisible(False)


        # 是否复原数据库中夹爪的角度
        cursor = db.cursor()
        sql = "UPDATE wyb_zp_temp SET zp_all_pos = %s WHERE id = 1"
        cursor.execute(sql, ('0&0&0&0&0&0',))
        db.commit()


    def system_wyb(self):

        self.zhujiemian = wyb()
        self.zhujiemian.ui.show()
        self.ui.close()


    def system_exit(self):

        self.ui.close()


    def system_test(self):

        self.begin_sys = robot_act_new.test_tongxun_connect()

        if self.begin_sys == 1:
            QMessageBox.information(QMessageBox(), '测试', '所有设备通讯成功！可以开始测量！')
            self.ui.system_wyb.setVisible(True)
            self.ui.system_digital.setVisible(True)
            self.ui.system_test.setVisible(False)
        elif self.begin_sys == -1:
            QMessageBox.information(QMessageBox(), '测试', '有设备通讯失败请检查')




class wyb():

    def __init__(self):


        self.ui = QUiLoader().load(addr_ui + "/wyb.ui")

        self.ui.wyb_equip.clicked.connect(self.wyb_equip)  # 下一步

        self.ui.wyb_start.clicked.connect(self.wyb_start)

        self.ui.wyb_manual.clicked.connect(self.wyb_manual)

        self.ui.wyb_renew.clicked.connect(self.wyb_renew)

        self.ui.wyb_find.clicked.connect(self.wyb_find)

        self.ui.wyb_report.clicked.connect(self.wyb_report)

        self.ui.wyb_new.clicked.connect(self.wyb_new)

        self.ui.wyb_rotate.clicked.connect(self.wyb_rotate)


        self.ui.wyb_exit.clicked.connect(self.wyb_exit)  # 下一步

        linshi_addr = "QLabel{border-image: url(" + addr_ui + "/登录底板.png)}"
        self.ui.wyb_pic.setStyleSheet(linshi_addr)

        self.ui.wyb_label.setStyleSheet("color: red;")
        self.ui.wyb_equip.setStyleSheet("background-color: yellow;")
        self.ui.wyb_manual.setStyleSheet("background-color: yellow;")
        self.ui.wyb_find.setStyleSheet("background-color: yellow;")


    def wyb_equip(self):

        self.zhujiemian = equip()
        self.zhujiemian.ui.show()
        self.ui.close()


    def wyb_start(self):

        self.zhujiemian = run()
        self.zhujiemian.ui.show()
        self.ui.close()


    def wyb_manual(self):

        cursor = db.cursor()
        sql = "SELECT num,nio FROM wyb_test_info WHERE equip_num = 1"
        cursor.execute(sql)
        db.commit()
        result = cursor.fetchone()
        if result is None:
            QMessageBox.information(self.ui, '运行', ' 未装夹，请装夹后再点击测量')
            return
        else:
            self.zhujiemian = run_manual()
            self.zhujiemian.ui.show()
            self.ui.close()


    def wyb_renew(self):

        self.zhujiemian = renew_wyb()
        self.zhujiemian.ui.show()
        self.ui.close()


    def wyb_find(self):

        self.zhujiemian = find_way()
        self.zhujiemian.ui.show()
        self.ui.close()


    def wyb_report(self):

        self.zhujiemian = report()
        self.zhujiemian.ui.show()
        self.ui.close()


    def wyb_new(self):

        self.zhujiemian = wyb_new()
        self.zhujiemian.ui.show()
        self.ui.close()


    def wyb_rotate(self):

        self.zhujiemian = rotate_zp_ai()
        self.zhujiemian.ui.show()
        self.ui.close()



    def wyb_exit(self):

        self.zhujiemian = system_wyb()
        self.zhujiemian.ui.show()
        self.ui.close()



class rotate_zp_ai(QWidget):


    def __init__(self):

        super(rotate_zp_ai, self).__init__()

        self.ui = QUiLoader().load(addr_ui + "/rotate_zp.ui")

        self.ui.rotate_zp_exit.clicked.connect(self.rotate_zp_exit)

        self.ui.rotate_zp_ok.clicked.connect(self.rotate_zp_ok)

        linshi_addr = "QLabel{border-image: url(" + addr_ui + "/登录底板.png)}"
        self.ui.rotate_zp_pic.setStyleSheet(linshi_addr)

        self.ui.rotate_zp_text.setText("请选择扩展功能！")

        self.ui.rotate_zp_text.setStyleSheet(
            '''color: red; justify-content: center; align-items: center; text-align: center;''')

        self.key = 0



    def rotate_zp_ok(self):

        if self.key == 0:

            self.key = 1

            self.ui.rotate_zp_text.setText("正在转转盘，不要进行任何操作！！等待提示！")

            rotate_zp_thread_name = threading.Thread(target=self.rotate_zp_thread)

            rotate_zp_thread_name.start()

        else:

            self.ui.rotate_zp_text.setText("请先等待转盘旋转完毕后，再旋转！")


    def rotate_zp_thread(self):

        zp_value = int(self.ui.rotate_zp_num.currentText())
        rotate_zp.send_zp_start(zp_value)


        self.ui.rotate_zp_text.setText("旋转完毕！")

        self.key = 0




    def rotate_zp_exit(self):


        if self.key == 0:

            self.zhujiemian = wyb()
            self.zhujiemian.ui.show()
            self.ui.close()

        else:

            self.ui.rotate_zp_text.setText("请先等待转盘旋转完毕后！再退出")

class wyb_new(QWidget):


    def __init__(self):

        super(wyb_new, self).__init__()

        self.ui = QUiLoader().load(addr_ui + "/wyb_new.ui")

        self.ui.wyb_new_ok.clicked.connect(self.wyb_new_ok)



        self.ui.wyb_new_exit.clicked.connect(self.wyb_new_exit)

        linshi_addr = "QLabel{border-image: url(" + addr_ui + "/登录底板.png)}"
        self.ui.wyb_new_pic.setStyleSheet(linshi_addr)





    def wyb_new_ok(self):


        cursor = db.cursor()


        sql1 = """
                                        SELECT 
                                            CASE 
                                                WHEN NOT EXISTS (
                                                    SELECT type FROM wyb_name_info WHERE type = %s
                                                )
                                                THEN -1
                                                ELSE 1
                                            END AS result
                                        """

        sql2 = """
                                            SELECT 
                                                CASE 
                                                    WHEN NOT EXISTS (
                                                        SELECT nio FROM wyb_name_info WHERE nio = %s
                                                    )
                                                    THEN -1
                                                    ELSE 1
                                                END AS result
                                            """



        # 一、导入 写好标定数据的 excel 表格 ！！！

        file_filter = "Excel files (*.xlsx)"
        options = QFileDialog.Options()
        default_path = "D:/new_wyb_sys/renew"
        # 弹出文件选择对话框，让用户选择文件
        path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", default_path, file_filter, options=options)

        if path:  # 确保用户选择了文件夹

            wb = load_workbook(path)

            # 选择活动工作表，默认为第一个工作表
            ws = wb.active



            # 二、先检查 nio 和 type 万用表编号和型号 ！！！！

            wyb_type = ws.cell(row=4, column=3).value
            wyb_nio = ws.cell(row=3, column=3).value

            cursor.execute(sql1, (wyb_type,))
            db.commit()
            result = cursor.fetchone()[0]

            if result == 1:

                QMessageBox.information(QMessageBox(), '录入', "请检查万用表型号，该型号已录入！")

            else:

                cursor.execute(sql2, (wyb_nio,))
                db.commit()
                result = cursor.fetchone()[0]

                if result == 1:

                    QMessageBox.information(QMessageBox(), '录入', "请检查nio，该nio已录入！")

                else:

                    # 三、导入数据入 wyb_name_info ！！！！！

                    wyb_num = wyb_type.replace('&', '') + '001'

                    sql2 = 'INSERT INTO wyb_name_info (`num`, `type`, `nio`) VALUES (%s, %s, %s) '
                    cursor.execute(sql2, (wyb_num, wyb_type, wyb_nio))
                    db.commit()



                    #  四、导入数据入 wyb_info ！！！！！

                    d_all_wyb = ws.cell(row=5, column=3).value
                    d_wyb_lcd = ws.cell(row=6, column=3).value
                    zp_pos = ws.cell(row=7, column=3).value
                    shun_ni = ws.cell(row=8, column=3).value.split('&')
                    shun_angle = shun_ni[0]
                    ni_angle = shun_ni[1]
                    dang_num = ws.cell(row=9, column=3).value

                    sql2 = 'INSERT INTO wyb_info (`type`,`dang_num`, `key_num`, `lcd_pos`, `zp_pos`, `key1_pos`, `key2_pos`, `k1_pos`' \
                           ', `k2_pos`, `k3_pos`, `k4_pos`, `shun_angle`, `ni_angle`, `d_all_wyb`, `d_wyb_lcd`) ' \
                           'VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s) '
                    cursor.execute(sql2, (wyb_type, dang_num, '0', '0', zp_pos, '0', '0', None, None, None, None, shun_angle, ni_angle, d_all_wyb, d_wyb_lcd ))
                    db.commit()



                    #  四、导入数据入 wyb_unit_info ！！！！！

                    dang_all_angle = ws.cell(row=10, column=3).value
                    angle_minus = (float(dang_all_angle)) / (float(dang_num))

                    # 第一个 OFF 档
                    id_num = ws.cell(row=14, column=1).value
                    dang_angle = ws.cell(row=14, column=2).value
                    dang_type = ws.cell(row=14, column=3).value

                    sql2 = 'INSERT INTO wyb_unit_info (`type`, `dang_num`, `dang_angle`, `key`, `dang_type`) VALUES (%s, %s, %s, %s, %s) '
                    cursor.execute(sql2, (wyb_type, id_num, dang_angle, '0', dang_type))
                    db.commit()


                    for i in range( 15, dang_num + 14 ):

                        id_num = ws.cell(row=i, column=1).value
                        dang_type = ws.cell(row=i, column=3).value
                        dang_angle = round((dang_angle + angle_minus), 3)

                        if dang_angle >= 360.000:

                            dang_angle = round((dang_angle - 360.000), 3)


                        sql2 = 'INSERT INTO wyb_unit_info (`type`, `dang_num`, `dang_angle`, `key`, `dang_type`) VALUES (%s, %s, %s, %s, %s) '
                        cursor.execute(sql2, (wyb_type, id_num, dang_angle, '0', dang_type))
                        db.commit()


                    id_num = 1

                    #  五、导入数据入 wyb_value ！！！！！

                    for i in range(15, dang_num + 14):

                        if ws.cell(row=i, column=4).value is not None :

                            measure_all = str(ws.cell(row=i, column=4).value).split('&')


                            for measure in measure_all:

                                if ws.cell(row=i, column=5).value != 0:
                                    unit_value = 0
                                    unit = ws.cell(row=i, column=3).value

                                if '-' in measure:
                                    unit_value = -1
                                    unit = ws.cell(row=i, column=3).value + '-'
                                elif ('-' not in measure) and (ws.cell(row=i, column=5).value == 0):
                                    unit_value = 1
                                    if ('R' not in str(ws.cell(row=i, column=3).value)):
                                        unit = ws.cell(row=i, column=3).value + '+'
                                    else:
                                        unit = ws.cell(row=i, column=3).value
                                        unit_value = 0

                                print(ws.cell(row=i, column=5).value)


                                hz = ws.cell(row=i, column=5).value

                                sql1 = 'INSERT INTO wyb_value (`id_num`, `num`, `dang_num`, `key`, `unit`, `hz`, `unit_value`,`measure`,' \
                                       ' `mea_time`, `m_1`, `m_2`, `m_3`, `m_average`,' \
                                       '`m_max`, `m_min`, `m_sz`, `m_xd`, `m_repeat`, `pass`, `ol`) ' \
                                       ' VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)'

                                dang_num = ws.cell(row=i, column=1).value

                                cursor.execute(sql1, (id_num, wyb_num, dang_num, 0, unit, hz, unit_value, measure, None, None
                                                      , None, None, None, None, None, None, None, None, None, None))

                                db.commit()

                                id_num = id_num + 1










            wb.save(path)

            QMessageBox.information(QMessageBox(), '录入', "录入成功！")



    def wyb_new_exit(self):

        self.zhujiemian = wyb()
        self.zhujiemian.ui.show()
        self.ui.close()



class report(QWidget):


    def __init__(self):

        super(report, self).__init__()

        self.ui = QUiLoader().load(addr_ui + "/report.ui")

        self.ui.report_find.clicked.connect(self.report_find)

        self.ui.report_export.clicked.connect(self.report_export)

        self.ui.report_exit.clicked.connect(self.report_exit)

        self.ui.report_nio.clear()



        # 填充 nio

        cursor = db.cursor()

        cursor.execute("SELECT nio FROM wyb_name_info")

        db.commit()

        for row in cursor.fetchall():
            # 假设查询结果是一个元组，并且我们只关心第一个元素
            nio = row[0]
            self.ui.report_nio.addItem(nio)  # 将结果添加到QComboBox





    def report_export(self):

        file_filter = "Excel files (*.xlsx)"
        options = QFileDialog.Options()

        # 弹出文件选择对话框，让用户选择文件
        path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", "", file_filter, options=options)

        if path:  # 确保用户选择了文件夹

            wb = load_workbook(path)

            # 选择活动工作表，默认为第一个工作表
            ws = wb.active

            wyb_nio = str(self.ui.report_nio.currentText())







            #  导出 日期，nio编号，型号

            cursor = db.cursor()

            sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
            cursor.execute(sql, (wyb_nio,))
            result = cursor.fetchone()[0]

            wyb_type = result.split('&')[1] if len(result.split('&')) > 1 else ''
            if wyb_type == 'VICTOR&VC890C+':
                wyb_type = 'VICTOR胜利'



            sql = "SELECT num FROM wyb_name_info WHERE nio = %s"
            cursor.execute(sql, (wyb_nio,))
            wyb_num = cursor.fetchone()

            sql = "SELECT mea_time FROM wyb_value WHERE num = %s"
            cursor.execute(sql, (wyb_num,))
            wyb_time = str(cursor.fetchall()[0][0])

            date_obj = datetime.strptime(wyb_time, '%Y%m%d%H%M%S')

            formatted_date = date_obj.strftime('%Y-%m-%d')


            ws.cell(row=2, column=5).value = str(wyb_type)
            ws.cell(row=2, column=7).value = str(wyb_nio)

            ws.cell(row=5, column=2).value = str(formatted_date)
            ws.cell(row=5, column=5).value = str(formatted_date)
            ws.cell(row=5, column=10).value = str(formatted_date)






            # 导出 测点 信息 ！！！

            for i in range(8, 100):

                unit1 = ws.cell(row=i, column=1).value

                if unit1 is not None:

                    unit2 = ws.cell(row=i, column=2).value

                    mea_value = ws.cell(row=i, column=3).value

                    mea_unit = str(self.report_find_unit(unit1,unit2))

                    if '-' in mea_unit:

                        mea_value = int(mea_value)
                        mea_value = str(- mea_value)

                    print(mea_unit)

                    sql = "SELECT m_1, m_2, m_3, m_average, m_sz, m_repeat, unit FROM wyb_value WHERE num = %s AND measure = %s AND unit LIKE %s"
                    cursor.execute(sql, ('VICTORVC890C+001', str(mea_value), '%' + mea_unit + '%'))
                    result = cursor.fetchone()


                    if result is None:

                        m_value = [-1, -1, -1, -1, -1, -1]

                    elif result[0] is None:

                        m_value = [-1, -1, -1, -1, -1, -1]

                    else:

                        m_value = list(result)

                    ws.cell(row=i, column=4).value = m_value[0]
                    ws.cell(row=i, column=5).value = m_value[1]
                    ws.cell(row=i, column=6).value = m_value[2]
                    ws.cell(row=i, column=7).value = m_value[3]
                    ws.cell(row=i, column=8).value = m_value[4]
                    ws.cell(row=i, column=9).value = m_value[5]






            # 保存工作簿
            wb.save(path)

            QMessageBox.information(QMessageBox(), '报告', "导出成功！")



    def report_find_unit(self,unit1,unit2):

        if '正' in unit1:

            if unit2 == 'mV':
                return 'MDCV+'
            elif unit2 == 'V':
                return 'DCV+'
            elif unit2 == 'uA':
                return 'UDCA+'
            elif unit2 == 'mA':
                return 'MDCA+'
            elif unit2 == 'A':
                return 'DCA+'

        elif '负' in unit1:

            if unit2 == 'mV':
                return 'MDCV-'
            elif unit2 == 'V':
                return 'DCV-'
            elif unit2 == 'uA':
                return 'UDCA-'
            elif unit2 == 'mA':
                return 'MDCA-'
            elif unit2 == 'A':
                return 'DCA-'

        elif '交流' in unit1:

            if unit2 == 'mV':
                return 'MACV'
            elif unit2 == 'V':
                return 'ACV'
            elif unit2 == 'uA':
                return 'UACA'
            elif unit2 == 'mA':
                return 'MACA'
            elif unit2 == 'A':
                return 'ACA'

        elif '电阻' in unit1:

            if unit2 == 'Ω':
                return 'R'
            elif unit2 == 'kΩ':
                return 'KR'
            elif unit2 == 'MΩ':
                return 'MR'



    def report_find(self):


        wyb_nio = str(self.ui.report_nio.currentText())

        cursor = db.cursor()

        cursor.execute("SELECT type FROM wyb_name_info WHERE nio = %s",(wyb_nio,))

        db.commit()

        wyb_type = str(cursor.fetchone()[0])

        self.ui.report_type.setText(wyb_type)




        cursor.execute("SELECT num FROM wyb_name_info WHERE nio = %s", (wyb_nio,))

        db.commit()

        wyb_num = str(cursor.fetchone()[0])




        cursor.execute("SELECT m_repeat FROM wyb_value WHERE num = %s", (wyb_num,))

        db.commit()


        for row in cursor.fetchall():


            if row is not None and row[0] is not None:

                m_repeat = row[0]

                if '%' in m_repeat:

                    key = 1

                else:

                    key = -1

                    break

            else:

                key = -1

                break




        if key == -1 :

            self.ui.report_status.setText('未完成所有档位测量！')

        else:

            self.ui.report_status.setText('全部完成，可以导出！')



    def report_exit(self):

        self.zhujiemian = wyb()
        self.zhujiemian.ui.show()
        self.ui.close()






class renew_wyb(QWidget):


    def __init__(self):

        super(renew_wyb, self).__init__()

        self.ui = QUiLoader().load(addr_ui + "/renew.ui")

        self.ui.renew_ok.clicked.connect(self.renew_ok)

        self.ui.renew_in.clicked.connect(self.renew_in)

        self.ui.renew_exit.clicked.connect(self.renew_exit)

        linshi_addr = "QLabel{border-image: url(" + addr_ui + "/登录底板.png)}"
        self.ui.renew_pic.setStyleSheet(linshi_addr)


    def renew_ok(self):

        wyb_nio = str(self.ui.renew_nio.text())
        wyb_type = str(self.ui.renew_type.text())

        cursor = db.cursor()

        sql1 = """
        SELECT 
            CASE 
                WHEN NOT EXISTS (
                    SELECT nio FROM wyb_name_info WHERE nio = %s
                )
                THEN -1
                ELSE 1
            END AS result
        """

        cursor.execute(sql1, (wyb_nio,))
        db.commit()

        result = cursor.fetchone()

        # 处理结果
        if result is not None:

            result_value = result[0]

            if result_value == -1:

                if '&' not in wyb_type:

                    QMessageBox.information(QMessageBox(), '录入', '型号输入错误!')

                else:

                    cursor.execute("SELECT type FROM wyb_name_info WHERE type = %s",(wyb_type,))
                    db.commit()

                    result = cursor.fetchone()

                    # 处理结果
                    if result is None:

                        QMessageBox.information(QMessageBox(), '录入', '该万用表型号为新增，需标定!')

                    else:

                        prefix = wyb_type.replace('&', '')
                        cursor.execute("SELECT num FROM wyb_name_info")
                        db.commit()

                        # 获取查询结果
                        nums = cursor.fetchall()

                        max_number = 0

                        for num_tuple in nums:
                            num_str = num_tuple[0]
                            # 检查字符串是否以特定前缀开始
                            if num_str.startswith(prefix):
                                # 从特定前缀后分割字符串，并尝试提取数字
                                # 去掉前缀
                                num_suffix = num_str[len(prefix):]
                                # 尝试转换为整数
                                num_digits = int(num_suffix)

                                # 更新最大数字
                                max_number = max(max_number, num_digits)

                        # 计算新编号
                        new_number = max_number + 1
                        # 构造新的num值，保持原有格式
                        wyb_num = f"{prefix}{new_number:03d}"
                        wyb_old_num = f"{prefix}{max_number:03d}"

                        sql2 = 'INSERT INTO wyb_name_info (`num`, `type`, `nio`) VALUES (%s, %s, %s) '
                        cursor.execute(sql2, (wyb_num, wyb_type, wyb_nio))
                        db.commit()

                        sql1 = """
                        INSERT INTO wyb_value (`id_num`, `num`, `dang_num`, `key`, `unit`, `hz`, `unit_value`, 
                                                      `measure`, `mea_time`, `m_1`, `m_2`, `m_3`, `m_average`, 
                                                      `m_max`, `m_min`, `m_sz`, `m_xd`, `m_repeat`, `pass`, `ol`)
                        SELECT `id_num`, %s AS `num`, `dang_num`, `key`, `unit`, `hz`, `unit_value`, 
                               `measure`, NULL AS `mea_time`, NULL AS `m_1`, NULL AS `m_2`, NULL AS `m_3`, 
                               NULL AS `m_average`, NULL AS `m_max`, NULL AS `m_min`, NULL AS `m_sz`, 
                               NULL AS `m_xd`, NULL AS `m_repeat`, NULL AS `pass`, NULL AS `ol`
                        FROM wyb_value
                        WHERE num = %s
                        """

                        cursor.execute(sql1, (wyb_num, wyb_old_num))

                        db.commit()

                        QMessageBox.information(QMessageBox(), '录入', '录入成功!')

            else:
                QMessageBox.information(QMessageBox(), '录入', '录入失败，该万用表nio型号已存在!')
        else:
            QMessageBox.information(QMessageBox(), '录入', '错误!')



    def renew_in(self):

        file_filter = "Excel files (*.xlsx)"
        options = QFileDialog.Options()
        default_path = "D:/new_wyb_sys/renew"
        # 弹出文件选择对话框，让用户选择文件
        path, _ = QFileDialog.getOpenFileName(self, "选择Excel文件", default_path, file_filter, options=options)

        if path:  # 确保用户选择了文件夹

            key = 0

            wb = load_workbook(path)

            # 选择活动工作表，默认为第一个工作表
            ws = wb.active

            text1 = ws.cell(row=20, column=1).value

            for i in range(2, 200):


                if (ws.cell(row=i, column=1).value) is not None:

                    wyb_nio = str(ws.cell(row=i, column=1).value)
                    wyb_type = str(ws.cell(row=i, column=2).value)

                    cursor = db.cursor()

                    sql1 = """
                    SELECT 
                        CASE 
                            WHEN NOT EXISTS (
                                SELECT nio FROM wyb_name_info WHERE nio = %s
                            )
                            THEN -1
                            ELSE 1
                        END AS result
                    """

                    cursor.execute(sql1, (wyb_nio,))
                    db.commit()

                    result = cursor.fetchone()

                    # 处理结果
                    if result is not None:

                        result_value = result[0]

                        if result_value == -1:

                            if '&' not in wyb_type:

                                pass

                            else:

                                cursor.execute("SELECT type FROM wyb_name_info WHERE type = %s",(wyb_type,))
                                db.commit()

                                result = cursor.fetchone()

                                # 处理结果
                                if result is None:

                                    key = 1

                                    pass

                                else:

                                    prefix = wyb_type.replace('&', '')
                                    cursor.execute("SELECT num FROM wyb_name_info")
                                    db.commit()

                                    # 获取查询结果
                                    nums = cursor.fetchall()

                                    max_number = 0

                                    for num_tuple in nums:
                                        num_str = num_tuple[0]
                                        # 检查字符串是否以特定前缀开始
                                        if num_str.startswith(prefix):
                                            # 从特定前缀后分割字符串，并尝试提取数字
                                            # 去掉前缀
                                            num_suffix = num_str[len(prefix):]
                                            # 尝试转换为整数
                                            num_digits = int(num_suffix)

                                            # 更新最大数字
                                            max_number = max(max_number, num_digits)

                                    # 计算新编号
                                    new_number = max_number + 1
                                    # 构造新的num值，保持原有格式
                                    wyb_num = f"{prefix}{new_number:03d}"
                                    wyb_old_num = f"{prefix}{max_number:03d}"

                                    sql2 = 'INSERT INTO wyb_name_info (`num`, `type`, `nio`) VALUES (%s, %s, %s) '
                                    cursor.execute(sql2, (wyb_num, wyb_type, wyb_nio))
                                    db.commit()

                                    sql1 = """
                                    INSERT INTO wyb_value (`id_num`, `num`, `dang_num`, `key`, `unit`, `hz`, `unit_value`, 
                                                                  `measure`, `mea_time`, `m_1`, `m_2`, `m_3`, `m_average`, 
                                                                  `m_max`, `m_min`, `m_sz`, `m_xd`, `m_repeat`, `pass`, `ol`)
                                    SELECT `id_num`, %s AS `num`, `dang_num`, `key`, `unit`, `hz`, `unit_value`, 
                                           `measure`, NULL AS `mea_time`, NULL AS `m_1`, NULL AS `m_2`, NULL AS `m_3`, 
                                           NULL AS `m_average`, NULL AS `m_max`, NULL AS `m_min`, NULL AS `m_sz`, 
                                           NULL AS `m_xd`, NULL AS `m_repeat`, NULL AS `pass`, NULL AS `ol`
                                    FROM wyb_value
                                    WHERE num = %s
                                    """

                                    cursor.execute(sql1, (wyb_num, wyb_old_num))

                                    db.commit()

                        else:

                            key = 2

                    else:

                        key = 3


            if key == 1:
                QMessageBox.information(QMessageBox(), '录入', "请检查表格的型号，存在 未标定 的型号！")
            elif key == 2:
                QMessageBox.information(QMessageBox(), '录入', "请检查表格的nio，存在 重复 或 之前录入过 的问题，不过不影响正常录入！")
            elif key == 3:
                QMessageBox.information(QMessageBox(), '录入', "数据库错误！")


            # 保存工作簿
            wb.save(path)

            QMessageBox.information(QMessageBox(), '录入', "录入成功！")


    def renew_exit(self):

        self.zhujiemian = wyb()
        self.zhujiemian.ui.show()
        self.ui.close()






class find_way():

    def __init__(self):

        self.ui = QUiLoader().load(addr_ui + "/find_way.ui")

        self.ui.find_nio.clicked.connect(self.find_nio)

        self.ui.find_type.clicked.connect(self.find_type)

        self.ui.find_type_day.clicked.connect(self.find_type_day)

        self.ui.find_day.clicked.connect(self.find_day)

        self.ui.find_exit.clicked.connect(self.find_exit)

        linshi_addr = "QLabel{border-image: url(" + addr_ui + "/登录底板.png)}"
        self.ui.find_way_pic.setStyleSheet(linshi_addr)

    def find_nio(self):

        global find_wyb_way

        find_wyb_way = 'nio'

        self.zhujiemian = find_result()
        self.zhujiemian.ui.show()
        self.ui.close()


    def find_type(self):

        global find_wyb_way

        find_wyb_way = 'type'

        self.zhujiemian = find_result()
        self.zhujiemian.ui.show()
        self.ui.close()


    def find_type_day(self):

        global find_wyb_way

        find_wyb_way = 'type_day'

        self.zhujiemian = find_result()
        self.zhujiemian.ui.show()
        self.ui.close()


    def find_day(self):

        global find_wyb_way

        find_wyb_way = 'day'

        self.zhujiemian = find_result()
        self.zhujiemian.ui.show()
        self.ui.close()


    def find_exit(self):

        self.zhujiemian = wyb()
        self.zhujiemian.ui.show()
        self.ui.close()






class equip():

    global vector_wyb

    def __init__(self):

        self.ui = QUiLoader().load(addr_ui + "/equip.ui")

        self.ui.equip_start.setStyleSheet("background-color: yellow;")
        self.ui.equip_ok.setStyleSheet("background-color: yellow;")
        self.ui.equip_next.setStyleSheet("background-color: yellow;")

        self.ui.equip_ok.clicked.connect(self.equip_ok)

        self.ui.equip_start.clicked.connect(self.equip_start)

        self.ui.equip_next.clicked.connect(self.equip_next)

        self.ui.equip_chai.clicked.connect(self.equip_chai)

        self.ui.equip_alter.clicked.connect(self.equip_alter)

        self.ui.equip_exit.clicked.connect(self.equip_exit)

        self.table = self.ui.equip_table

        self.equip_graph = self.ui.equip_graph

        self.ui.equip_start.setVisible(True)
        self.ui.equip_help.setStyleSheet("color: red;")

        vector_wyb.clear()

        # 初始化 装夹 表格
        cursor = db.cursor()

        cursor.execute("SELECT MAX(equip_num) FROM wyb_test_info")
        max_value = cursor.fetchone()[0]
        print(max_value)

        if max_value is not None:

            max_value = int(max_value)

            print(max_value)
            # 步骤3: 从1遍历到a
            for equip_num_py in range(1, max_value + 1):
                # 步骤4: 查询num和nio字段的值
                cursor.execute("SELECT nio FROM wyb_test_info WHERE equip_num = %s", (equip_num_py,))
                result1 = cursor.fetchone()
                cursor.execute("SELECT type FROM wyb_name_info WHERE nio = %s", (result1,))
                result2 = cursor.fetchone()

                # 步骤5: 将结果添加到vector_wyb数组中
                if result1 and result2:
                    vector_wyb.append([equip_num_py, result1[0], result2[0]])

        self.table_length = len(vector_wyb)


        if self.table_length != 0:

            for num in range(1, self.table_length + 1):

                newItem = QTableWidgetItem(str(vector_wyb[num-1][1]))
                self.table.setItem(num, 1, newItem)

                newItem1 = QTableWidgetItem(vector_wyb[num-1][2])
                self.table.setItem(num, 2, newItem1)




        # 初始化 图像

        self.alter_image()

        self.mode = -1

        linshi_addr = "QLabel{border-image: url(" + addr_ui + "/登录底板.png)}"
        self.ui.equip_pic.setStyleSheet(linshi_addr)

    def equip_photo(self, ph_name):

        global equip_image

        time.sleep(3)
        testGrabImage.paizhao(ph_name)
        photo_name = "E:/photo/" + ph_name + ".jpg"
        self.alter_image()
        time.sleep(2)
        dedaozhi = ewm_and_lcd.erweima_shibei(photo_name)

        equip_image = 'E:/photo/ewm_status.jpg'
        self.alter_image()

        if dedaozhi != 'shibai':
            return dedaozhi
        elif dedaozhi == 'shibai':
            time.sleep(3)
            testGrabImage.paizhao(ph_name)
            time.sleep(2)
            dedaozhi = ewm_and_lcd.erweima_shibei(photo_name, 105)
            self.alter_image()
            if dedaozhi == 'shibai':
                time.sleep(3)
                testGrabImage.paizhao(ph_name)
                time.sleep(2)
                dedaozhi = ewm_and_lcd.erweima_shibei(photo_name, 130)
                self.alter_image()
                if dedaozhi == 'shibai':
                    time.sleep(3)
                    testGrabImage.paizhao(ph_name)
                    time.sleep(2)
                    dedaozhi = ewm_and_lcd.erweima_shibei(photo_name, 140)
                    self.alter_image()
                    if dedaozhi == 'shibai':
                        time.sleep(3)
                        testGrabImage.paizhao(ph_name)
                        time.sleep(2)
                        dedaozhi = ewm_and_lcd.erweima_shibei(photo_name, 95)
                        self.alter_image()
                        if dedaozhi == 'shibai':
                            time.sleep(3)
                            testGrabImage.paizhao(ph_name)
                            time.sleep(2)
                            dedaozhi = ewm_and_lcd.erweima_shibei(photo_name, 120)
                            self.alter_image()
                            if dedaozhi == 'shibai':
                                time.sleep(3)
                                testGrabImage.paizhao(ph_name)
                                time.sleep(2)
                                dedaozhi = ewm_and_lcd.erweima_shibei(photo_name, 100)
                                self.alter_image()
                                if dedaozhi == 'shibai':
                                    time.sleep(3)
                                    testGrabImage.paizhao(ph_name)
                                    time.sleep(2)
                                    dedaozhi = ewm_and_lcd.erweima_shibei(photo_name, 90)
                                    self.alter_image()

            return dedaozhi


    def equip_ok(self):

        reply = QMessageBox.question(QMessageBox(), '装夹', '已全部完成人工装夹，请将 双门 关闭，双门是否已关闭？',
                                     QMessageBox.Yes | QMessageBox.No | QMessageBox.Yes)

        if reply == QMessageBox.Yes:
            self.zhujiemian = wyb()
            self.zhujiemian.ui.show()
            self.ui.close()
        else:
            QMessageBox.information(QMessageBox(), '装夹', '请将 双门 关闭!')
            return


    def equip_start(self):

        global vector_wyb,equip_image

        self.mode = 1

        reply = QMessageBox.question(QMessageBox(), '装夹', '是否开始装夹( 将清空数据 )', QMessageBox.Yes | QMessageBox.No | QMessageBox.Yes)

        if reply == QMessageBox.Yes:

            self.ui.equip_start.setVisible(False)

            # 清空
            self.table.setItem(1, 1, None)
            self.table.setItem(1, 2, None)
            self.table.setItem(2, 1, None)
            self.table.setItem(2, 2, None)
            self.table.setItem(3, 1, None)
            self.table.setItem(3, 2, None)
            self.table.setItem(4, 1, None)
            self.table.setItem(4, 2, None)
            self.table.setItem(5, 1, None)
            self.table.setItem(5, 2, None)
            self.table.setItem(6, 1, None)
            self.table.setItem(6, 2, None)
            self.table.setItem(7, 1, None)
            self.table.setItem(7, 2, None)
            self.table.setItem(8, 1, None)
            self.table.setItem(8, 2, None)

            vector_wyb.clear()


            # 数据库清空
            cursor = db.cursor()


            sql = "SELECT max(id) FROM wyb_test_info"
            cursor.execute(sql)
            db.commit()
            result = cursor.fetchone()
            max_id = int(result[0])
            print(max_id)
            for i in range(1, max_id + 1):

                sql1 = "UPDATE wyb_test_info SET num = NULL,nio = NULL,equip_num = NULL,dang_num = NULL WHERE id = %s"
                cursor.execute(sql1, (str(i),))
                db.commit()

            self.ui.equip_help.setText('机器臂正移动到装夹位置')


            # 机器人
            robot_act_new.movej_PTP_getJ(3, 60.0, robot_act_new.P_begin)

            self.ui.equip_help.setText('机器臂移动完毕')



            # 扫码出图

            dedaozhi = self.equip_photo('last_equip_ewm')

            if dedaozhi == 'shibai':

                QMessageBox.information(QMessageBox(), '装夹', "扫码失败，请重启程序，重新进入装夹界面！")

                return


            equip_num = int(dedaozhi)

            #equip_image = 'E:/photo/last_equip_ewm.jpg'
            #self.alter_image()

            if equip_num != 1:

                # 消息
                #message = "扫码完成，当前装夹位置为 {} 号 装夹位！".format(equip_num)
                #QMessageBox.information(QMessageBox(), '装夹', message)
                time.sleep(5)
                # 旋转转盘
                robot_act_new.send_zp_start(9 - equip_num)
                time.sleep(5)
                # 等待指令
                #QMessageBox.information(QMessageBox(), '装夹', "OK")

            elif equip_num == 1:

                # 消息
                message = "扫码完成，当前装夹位置为 1 号 装夹位！"
                QMessageBox.information(QMessageBox(), '装夹', message)



            # 再次确认扫码
            dedaozhi = self.equip_photo('begin_equip_ewm')

            if dedaozhi == 'shibai':

                QMessageBox.information(QMessageBox(), '装夹', "扫码失败，请重启程序，重新进入装夹界面！")

                return


            equip_num_again = int(dedaozhi)
            


            #equip_image = 'E:/photo/begin_equip_ewm.jpg'
            #self.alter_image()


            # 位置信息提示
            reply = QMessageBox.question(QMessageBox(), '装夹', '当前装夹位置为 {} 号 装夹位，请选择是否手动输入万用表编号？'.format(equip_num_again),
                                         QMessageBox.Yes | QMessageBox.No | QMessageBox.Yes)

            if reply == QMessageBox.Yes:

                text, ok_pressed = QInputDialog.getText(None, '装夹', '请输入万用表编号：')

                if not ok_pressed:
                    self.ui.equip_start.setVisible(True)
                    return

                text = str(text)
                wyb_num = text

                sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                cursor.execute(sql, (text,))
                db.commit()
                result = cursor.fetchone()

                # 防错
                while result is None:

                    message = "该万用表nio型号为空，请重新输入！"
                    QMessageBox.information(QMessageBox(), '装夹', message)

                    text, ok_pressed = QInputDialog.getText(None, '装夹', '请输入万用表编号：')

                    if not ok_pressed:
                        self.ui.equip_start.setVisible(True)
                        return

                    text = str(text)
                    wyb_num = text

                    sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                    cursor.execute(sql, (text,))
                    db.commit()
                    result = cursor.fetchone()


            elif reply == QMessageBox.No:

                text, ok_pressed = QInputDialog.getText(None, '装夹', '请输入万用表编号：')

                if not ok_pressed:
                    self.ui.equip_start.setVisible(True)
                    return

                text = str(text)
                wyb_num = text

                sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                cursor.execute(sql, (text,))
                db.commit()
                result = cursor.fetchone()

                # 防错
                while result is None:

                    message = "该万用表nio型号为空，请重新输入！"
                    QMessageBox.information(QMessageBox(), '装夹', message)

                    text, ok_pressed = QInputDialog.getText(None, '装夹', '请输入万用表编号：')

                    if not ok_pressed:
                        self.ui.equip_start.setVisible(True)
                        return

                    text = str(text)
                    wyb_num = text

                    sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                    cursor.execute(sql, (text,))
                    db.commit()
                    result = cursor.fetchone()

                '''
                message = "当前装夹位置为 {} 号 装夹位，请将万用表二维码放到摄像头下！".format(equip_num_again)
                QMessageBox.information(QMessageBox(), '装夹', message)


                # 扫码出图
                photo_name = "begin_wyb_ewm"
                testGrabImage.paizhao(photo_name)
                equip_image = 'E:/photo/begin_wyb_ewm.jpg'
                self.alter_image()
                time.sleep(2)
                dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg")
                wyb_num = str(dedaozhi)
                equip_image = 'E:/photo/begin_wyb_ewm.jpg'
                self.alter_image()
                print(wyb_num)


                sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                cursor.execute(sql, (wyb_num,))
                db.commit()
                result = cursor.fetchone()

                # 防错
                while result is None:

                    message = "该万用表nio型号为空，请检查是否录入，重新将录入的万用表二维码放到摄像头下！"
                    QMessageBox.information(QMessageBox(), '装夹', message)

                    photo_name = "begin_wyb_ewm"
                    testGrabImage.paizhao(photo_name)
                    equip_image = 'E:/photo/begin_wyb_ewm.jpg'
                    self.alter_image()
                    time.sleep(2)
                    dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg")
                    wyb_num = str(dedaozhi)
                    equip_image = 'E:/photo/begin_wyb_ewm.jpg'
                    self.alter_image()
                    sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                    cursor.execute(sql, (wyb_num,))
                    db.commit()
                    result = cursor.fetchone()
                '''

            elif (reply != QMessageBox.Yes) and (reply != QMessageBox.No):

                self.ui.equip_start.setVisible(True)
                return

            wyb_style = str(result[0])

            sql = "SELECT dang_num FROM wyb_info WHERE type = %s"
            cursor.execute(sql, (wyb_style,))
            db.commit()
            result = cursor.fetchone()
            dang_num = int(result[0])

            sql = "SELECT num FROM wyb_name_info WHERE nio = %s"
            cursor.execute(sql, (wyb_num,))
            db.commit()
            result = cursor.fetchone()
            wyb_all_num = str(result[0])


            print(dang_num)
            print(wyb_all_num)

            for i in range( 1, dang_num + 1 ):

                sql1 = "UPDATE wyb_test_info SET num = %s WHERE id = %s"
                sql2 = "UPDATE wyb_test_info SET nio = %s WHERE id = %s"
                sql3 = "UPDATE wyb_test_info SET dang_num = %s WHERE id = %s"
                sql4 = "UPDATE wyb_test_info SET equip_num = %s WHERE id = %s"
                print(i)
                cursor.execute(sql1, (wyb_all_num, str(i)))
                db.commit()
                cursor.execute(sql2, (wyb_num, str(i)))
                db.commit()
                cursor.execute(sql3, (str(i), str(i)))
                db.commit()
                cursor.execute(sql4, (str(equip_num_again), str(i)))
                db.commit()





            # 万用表提示
            message = "已将 {} 号万用表 配于 1 号 装夹位，请开始装夹".format(wyb_num)
            QMessageBox.information(QMessageBox(), '装夹', message)
            QMessageBox.information(QMessageBox(), '装夹', ' 待安装完毕后 !!! 点击 下一台 按钮!!！')

            '''
            cursor = db.cursor()
            sql = "UPDATE wyb_equip SET equip_num = %d WHERE id = 1"
            cursor.execute(sql, (1,))
            db.commit()

            sql = "UPDATE wyb_equip SET wyb_style = %s WHERE id = 1"
            cursor.execute(sql, (wyb_style,))
            db.commit()
            '''

            newItem = QTableWidgetItem(wyb_num)  # 创建表格项---文本项目
            self.table.setItem(1, 1, newItem)  # 给指定单元格设置数据

            '''
            item = self.table.item(1, 0)  # 返回指定单元格的对象
            item.setTextAlignment(Qt.AlignRight | Qt.AlignBottom)'''

            newItem1 = QTableWidgetItem(wyb_style)  # 创建表格项---文本项目
            self.table.setItem(1, 2, newItem1)  # 给指定单元格设置数据

            vector_wyb.append([1,wyb_num,wyb_style])

        else:

            return


    def equip_chai(self):

        global vector_wyb

        self.mode = 2

        reply = QMessageBox.question(QMessageBox(), '拆卸', '是否开始拆卸？', QMessageBox.Yes | QMessageBox.No | QMessageBox.Yes)

        if reply == QMessageBox.Yes:

            # 清空

            self.table.setItem(1, 1, None)
            self.table.setItem(1, 2, None)
            self.table.setItem(2, 1, None)
            self.table.setItem(2, 2, None)
            self.table.setItem(3, 1, None)
            self.table.setItem(3, 2, None)
            self.table.setItem(4, 1, None)
            self.table.setItem(4, 2, None)
            self.table.setItem(5, 1, None)
            self.table.setItem(5, 2, None)
            self.table.setItem(6, 1, None)
            self.table.setItem(6, 2, None)
            self.table.setItem(7, 1, None)
            self.table.setItem(7, 2, None)
            self.table.setItem(8, 1, None)
            self.table.setItem(8, 2, None)

            vector_wyb.clear()


            # 数据库清空
            cursor = db.cursor()
            sql = "SELECT max(id) FROM wyb_test_info"
            cursor.execute(sql)
            db.commit()
            result = cursor.fetchone()
            max_id = int(result[0])
            print(max_id)
            for i in range(1, max_id + 1):

                sql1 = "UPDATE wyb_test_info SET num = NULL,nio = NULL,equip_num = NULL,dang_num = NULL WHERE id = %s"
                cursor.execute(sql1, (str(i), ))
                db.commit()



            QMessageBox.information(QMessageBox(), '拆卸', "请拆除该台万用表！然后点击下一步！")


        else:

            return


    def equip_next(self):

        global hang,wyb_style_all,wyb_num_all,equip_image



        if self.mode == 1:

            reply = QMessageBox.question(QMessageBox(), '装夹', '是否开始装夹 下一台 万用表？',
                                         QMessageBox.Yes | QMessageBox.No | QMessageBox.Yes)

            if reply == QMessageBox.Yes:

                if hang >= 9:

                    QMessageBox.critical(QMessageBox(), '警告', '当前所有装夹位已满！')
                    return

                else:

                    # 旋转转盘
                    robot_act_new.send_zp_start(1)

                    # 等待指令
                    #QMessageBox.information(QMessageBox(), '装夹', "OK")
                    time.sleep(5)

                    # 扫码出图
                    dedaozhi = self.equip_photo('next_equip_ewm')


                    if dedaozhi == 'shibai':


                        QMessageBox.information(QMessageBox(), '装夹', "扫码失败，请重启程序，重新进入装夹界面！")

                        return

                    equip_num_next = int(dedaozhi)


                    #equip_image = 'E:/photo/next_equip_ewm.jpg'
                    #self.alter_image()



                    # 位置信息提示
                    cursor = db.cursor()

                    reply = QMessageBox.question(QMessageBox(), '装夹', '当前装夹位置为 {} 号 装夹位，请选择是否手动输入万用表编号？'.format(equip_num_next),
                                                 QMessageBox.Yes | QMessageBox.No | QMessageBox.Yes)

                    if reply == QMessageBox.Yes:

                        text, ok_pressed = QInputDialog.getText(None, '装夹', '请输入万用表编号：')

                        if not ok_pressed:
                            self.ui.equip_start.setVisible(True)
                            return

                        text = str(text)
                        wyb_num = text
                        sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                        cursor.execute(sql, (text,))
                        db.commit()
                        result = cursor.fetchone()

                        # 防错
                        while result is None:
                            message = "该万用表nio型号为空，请重新输入！"
                            QMessageBox.information(QMessageBox(), '装夹', message)

                            text, ok_pressed = QInputDialog.getText(None, '装夹', '请输入万用表编号：')

                            if not ok_pressed:
                                self.ui.equip_start.setVisible(True)
                                return

                            text = str(text)
                            wyb_num = text

                            sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                            cursor.execute(sql, (text,))
                            db.commit()
                            result = cursor.fetchone()

                    elif reply == QMessageBox.No:

                        message = "当前装夹位置为 {} 号 装夹位，请将万用表二维码放到摄像头下！".format(equip_num_next)
                        QMessageBox.information(QMessageBox(), '装夹', message)

                        # 扫码出图
                        photo_name = "next_wyb_ewm"
                        testGrabImage.paizhao(photo_name)
                        equip_image = 'E:/photo/next_wyb_ewm.jpg'
                        self.alter_image()
                        time.sleep(2)
                        dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg")
                        wyb_num = str(dedaozhi)
                        equip_image = 'E:/photo/next_wyb_ewm.jpg'
                        self.alter_image()
                        print(wyb_num)

                        sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                        cursor.execute(sql, (wyb_num,))
                        db.commit()
                        result = cursor.fetchone()

                        # 防错
                        while result is None:
                            message = "该万用表nio型号为空，请检查，重新将万用表二维码放到摄像头下！"
                            QMessageBox.information(QMessageBox(), '装夹', message)

                            photo_name = "next_wyb_ewm"
                            testGrabImage.paizhao(photo_name)
                            equip_image = 'E:/photo/next_wyb_ewm.jpg'
                            self.alter_image()
                            time.sleep(2)
                            dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg")
                            wyb_num = str(dedaozhi)
                            equip_image = 'E:/photo/next_wyb_ewm.jpg'
                            self.alter_image()
                            sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                            cursor.execute(sql, (wyb_num,))
                            db.commit()
                            result = cursor.fetchone()

                    elif (reply != QMessageBox.Yes) and (reply != QMessageBox.No):

                        self.ui.equip_start.setVisible(True)
                        return


                    wyb_style = str(result[0])


                    sql = "SELECT dang_num FROM wyb_info WHERE type = %s"
                    cursor.execute(sql, (wyb_style,))
                    db.commit()
                    result = cursor.fetchone()
                    dang_num = int(result[0])


                    sql = "SELECT num FROM wyb_name_info WHERE nio = %s"
                    cursor.execute(sql, (wyb_num,))
                    db.commit()
                    result = cursor.fetchone()
                    wyb_all_num = str(result[0])



                    sql = "SELECT max(id) FROM wyb_test_info WHERE nio IS NOT NULL"
                    cursor.execute(sql)
                    db.commit()
                    result = cursor.fetchone()
                    max_id = int(result[0])
                    print(max_id)

                    for (i,j) in zip(range(max_id + 1, dang_num + max_id + 1), range(1, dang_num + 1)):


                        sql1 = "UPDATE wyb_test_info SET num = %s WHERE id = %s"
                        sql2 = "UPDATE wyb_test_info SET nio = %s WHERE id = %s"
                        sql3 = "UPDATE wyb_test_info SET dang_num = %s WHERE id = %s"
                        sql4 = "UPDATE wyb_test_info SET equip_num = %s WHERE id = %s"

                        cursor.execute(sql1, (wyb_all_num, str(i)))
                        db.commit()
                        cursor.execute(sql2, (wyb_num, str(i)))
                        db.commit()
                        cursor.execute(sql3, (str(j), str(i)))
                        db.commit()
                        cursor.execute(sql4, (str(equip_num_next), str(i)))
                        db.commit()



                    # 万用表提示
                    message = "已将 {} 该型号万用表 配对于 {} 号 装夹位，请开始装夹".format(wyb_style, equip_num_next)
                    QMessageBox.information(QMessageBox(), '装夹', message)
                    QMessageBox.information(QMessageBox(), '装夹', ' 待安装完毕后！！！ 请点击  下一台！！！ <br>若 全部 装夹完毕，请点击 装夹完毕！！<br> 若需要更换万用表装夹，请点击 修改装夹！')



                    # 表格填充对应信息
                    row = hang
                    column = 1

                    newItem = QTableWidgetItem(wyb_num)  # 创建表格项---文本项目
                    self.table.setItem(row, column, newItem)  # 给指定单元格设置数据

                    newItem = QTableWidgetItem(wyb_style)  # 创建表格项---文本项目
                    self.table.setItem(row, column + 1, newItem)  # 给指定单元格设置数据

                    hang = hang + 1

                    vector_wyb.append([hang,wyb_num,wyb_style])



            else:

                return

        elif self.mode == 2:

            reply = QMessageBox.question(QMessageBox(), '拆卸', '是否确定拆卸 下一台 万用表？', QMessageBox.Yes | QMessageBox.No | QMessageBox.Yes)

            if reply == QMessageBox.Yes:

                # 旋转转盘
                robot_act_new.send_zp_start(1)

                # 提示
                QMessageBox.information(QMessageBox(), '拆卸', '请拆卸！然后点击下一台')

            else:

                return

        else:

            # 提示
            QMessageBox.information(QMessageBox(), '装配', '请先点击 装夹 或 拆卸 指令后 再点击下一台')


    def equip_alter(self):

        global vector_wyb,equip_image

        text, ok_pressed = QInputDialog.getText(None, '更改', '请输入要调整的装夹位：')

        if ok_pressed:

            if text != '':
                # 机器人
                self.ui.equip_help.setText('机器臂正移动至初始装夹位置')
                robot_act_new.movej_PTP_getJ(3, 60.0, robot_act_new.P_begin)
                self.ui.equip_help.setText('机器臂移动完毕')


                # 扫码

                dedaozhi = self.equip_photo('alter_equip_ewm')

                if dedaozhi == 'shibai':
                    QMessageBox.information(QMessageBox(), '装夹', "扫码失败，请重启程序，重新进入装夹界面！")

                    return

                now_num = int(dedaozhi)


                #equip_image = 'E:/photo/alter_equip_ewm.jpg'
                #self.alter_image()

                # 消息
                #message = "当前装夹位置为 {} 号 装夹位,开始旋转！".format(now_num)
                #QMessageBox.information(QMessageBox(), '装夹', message)


                # 计算
                end_pos = int(text)

                key1 = 8 - now_num + end_pos

                if key1 == 8:

                    pass

                elif key1 > 8:

                    key1 = key1 - 8
                    robot_act_new.send_zp_start(int(key1))

                    # 等待指令
                    #QMessageBox.information(QMessageBox(), '装夹', "OK")

                elif key1 < 8:

                    robot_act_new.send_zp_start(int(key1))

                    # 等待指令
                    #QMessageBox.information(QMessageBox(), '装夹', "OK")


                time.sleep(2)

                # 再次扫码
                dedaozhi = self.equip_photo('do_equip_ewm')

                if dedaozhi == 'shibai':
                    QMessageBox.information(QMessageBox(), '装夹', "扫码失败，请重启程序，重新进入装夹界面！")

                    return

                end_pos = int(dedaozhi)


                #equip_image = 'E:/photo/do_equip_ewm.jpg'
                #self.alter_image()



                reply = QMessageBox.question(QMessageBox(), '装夹', '当前装夹位置为 {} 号 装夹位，请选择是否手动输入万用表编号？'.format(end_pos),
                                             QMessageBox.Yes | QMessageBox.No | QMessageBox.Yes)

                cursor = db.cursor()

                if reply == QMessageBox.Yes:

                    text, ok_pressed = QInputDialog.getText(None, '装夹', '请输入万用表编号：')

                    if not ok_pressed:
                        self.ui.equip_start.setVisible(True)
                        return

                    text = str(text)
                    wyb_num = text


                    sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                    cursor.execute(sql, (text,))
                    db.commit()
                    result = cursor.fetchone()

                    # 防错
                    while result is None:
                        message = "该万用表nio型号为空，请重新输入！"
                        QMessageBox.information(QMessageBox(), '装夹', message)

                        text, ok_pressed = QInputDialog.getText(None, '装夹', '请输入万用表编号：')
                        if not ok_pressed:
                            self.ui.equip_start.setVisible(True)
                            return
                        text = str(text)
                        wyb_num = text

                        sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                        cursor.execute(sql, (text,))
                        db.commit()
                        result = cursor.fetchone()


                elif reply == QMessageBox.No:

                    text, ok_pressed = QInputDialog.getText(None, '装夹', '请输入万用表编号：')

                    if not ok_pressed:
                        self.ui.equip_start.setVisible(True)
                        return

                    text = str(text)
                    wyb_num = text

                    sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                    cursor.execute(sql, (text,))
                    db.commit()
                    result = cursor.fetchone()

                    # 防错
                    while result is None:

                        message = "该万用表nio型号为空，请重新输入！"
                        QMessageBox.information(QMessageBox(), '装夹', message)

                        text, ok_pressed = QInputDialog.getText(None, '装夹', '请输入万用表编号：')

                        if not ok_pressed:
                            self.ui.equip_start.setVisible(True)
                            return

                        text = str(text)
                        wyb_num = text

                        sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                        cursor.execute(sql, (text,))
                        db.commit()
                        result = cursor.fetchone()

                    '''
                    message = "当前装夹位置为 {} 号 装夹位，请将万用表二维码放到摄像头下！".format(end_pos)
                    QMessageBox.information(QMessageBox(), '装夹', message)

                    # 扫码出图
                    photo_name = "alter_wyb_ewm"
                    testGrabImage.paizhao(photo_name)
                    equip_image = 'E:/photo/alter_wyb_ewm.jpg'
                    self.alter_image()
                    time.sleep(2)
                    dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg")
                    wyb_num = str(dedaozhi)
                    equip_image = 'E:/photo/alter_wyb_ewm.jpg'
                    self.alter_image()


                    sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                    cursor.execute(sql, (wyb_num,))
                    db.commit()
                    result = cursor.fetchone()

                    # 防错
                    while result is None:
                        message = "该万用表nio型号为空，请检查是否录入，重新将录入的万用表二维码放到摄像头下！"
                        QMessageBox.information(QMessageBox(), '装夹', message)

                        photo_name = "alter_wyb_ewm"
                        testGrabImage.paizhao(photo_name)
                        equip_image = 'E:/photo/alter_wyb_ewm.jpg'
                        self.alter_image()
                        time.sleep(2)
                        dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg")
                        wyb_num = str(dedaozhi)
                        equip_image = 'E:/photo/alter_wyb_ewm.jpg'
                        self.alter_image()

                        sql = "SELECT type FROM wyb_name_info WHERE nio = %s"
                        cursor.execute(sql, (wyb_num,))
                        db.commit()
                        result = cursor.fetchone()
                    '''

                elif (reply != QMessageBox.Yes) and (reply != QMessageBox.No):

                    self.ui.equip_start.setVisible(True)
                    return


                if end_pos > len(vector_wyb):


                    wyb_style = str(result[0])   # 万用表 型号

                    sql = "SELECT num FROM wyb_name_info WHERE nio = %s"
                    cursor.execute(sql, (wyb_num,))
                    db.commit()
                    result = cursor.fetchone()
                    wyb_all_num = str(result[0])

                    #  wyb_num     ：NIO 编号
                    #  wyb_all_num ：内码
                    #  wyb_style   ：万用表型号


                    sql = "SELECT dang_num FROM wyb_info WHERE type = %s"
                    cursor.execute(sql, (wyb_style,))
                    db.commit()
                    result = cursor.fetchone()

                    dang_num = int(result[0])



                    sql = "SELECT max(id) FROM wyb_test_info WHERE nio IS NOT NULL"
                    cursor.execute(sql)
                    db.commit()
                    result = cursor.fetchone()
                    max_id = int(result[0])


                    for (i, j) in zip(range(max_id + 1, dang_num + max_id + 1), range(1, dang_num + 1)):

                        sql1 = "UPDATE wyb_test_info SET num = %s WHERE id = %s"
                        sql2 = "UPDATE wyb_test_info SET nio = %s WHERE id = %s"
                        sql3 = "UPDATE wyb_test_info SET dang_num = %s WHERE id = %s"
                        sql4 = "UPDATE wyb_test_info SET equip_num = %s WHERE id = %s"

                        cursor.execute(sql1, (wyb_all_num, str(i)))
                        db.commit()
                        cursor.execute(sql2, (wyb_num, str(i)))
                        db.commit()
                        cursor.execute(sql3, (str(j), str(i)))
                        db.commit()
                        cursor.execute(sql4, (str(end_pos), str(i)))
                        db.commit()

                    sql6 = "SET @auto_id = 0"
                    sql7 = "UPDATE wyb_test_info SET id = (@auto_id := @auto_id + 1)"
                    sql8 = "ALTER TABLE wyb_test_info AUTO_INCREMENT = 1"

                    cursor.execute(sql6)
                    db.commit()
                    cursor.execute(sql7)
                    db.commit()
                    cursor.execute(sql8)
                    db.commit()

                    vector_wyb.append([end_pos, wyb_num, wyb_style])

                elif end_pos == len(vector_wyb):

                    wyb_style = str(result[0])

                    sql = "SELECT num FROM wyb_name_info WHERE nio = %s"
                    cursor.execute(sql, (wyb_num,))
                    db.commit()
                    result = cursor.fetchone()
                    wyb_all_num = str(result[0])

                    #  wyb_num     ：NIO 编号
                    #  wyb_all_num ：内码
                    #  wyb_style   ：万用表型号


                    sql = "SELECT dang_num FROM wyb_info WHERE type = %s"
                    cursor.execute(sql, (wyb_style,))
                    db.commit()
                    result = cursor.fetchone()
                    dang_num = int(result[0])


                    # 清空原来位置的万用表数据

                    sql3 = "UPDATE wyb_test_info SET dang_num = NULL,nio = NULL,equip_num = NULL,num = NULL WHERE equip_num = %s"

                    cursor.execute(sql3, (end_pos,))
                    db.commit()

                    sql = "SELECT max(id) FROM wyb_test_info WHERE nio IS NOT NULL"
                    cursor.execute(sql)
                    db.commit()
                    result = cursor.fetchone()
                    max_id = int(result[0])

                    for (i, j) in zip(range(max_id + 1, dang_num + max_id + 1), range(1, dang_num + 1)):
                        sql1 = "UPDATE wyb_test_info SET num = %s WHERE id = %s"
                        sql2 = "UPDATE wyb_test_info SET nio = %s WHERE id = %s"
                        sql3 = "UPDATE wyb_test_info SET dang_num = %s WHERE id = %s"
                        sql4 = "UPDATE wyb_test_info SET equip_num = %s WHERE id = %s"

                        cursor.execute(sql1, (wyb_all_num, str(i)))
                        db.commit()
                        cursor.execute(sql2, (wyb_num, str(i)))
                        db.commit()
                        cursor.execute(sql3, (str(j), str(i)))
                        db.commit()
                        cursor.execute(sql4, (str(end_pos), str(i)))
                        db.commit()

                    sql6 = "SET @auto_id = 0"
                    sql7 = "UPDATE wyb_test_info SET id = (@auto_id := @auto_id + 1)"
                    sql8 = "ALTER TABLE wyb_test_info AUTO_INCREMENT = 1"

                    cursor.execute(sql6)
                    db.commit()
                    cursor.execute(sql7)
                    db.commit()
                    cursor.execute(sql8)
                    db.commit()

                    vector_wyb[end_pos - 1] = [end_pos, wyb_num, wyb_style]

                else:

                    wyb_style = str(result[0])

                    sql = "SELECT num FROM wyb_name_info WHERE nio = %s"
                    cursor.execute(sql, (wyb_num,))
                    db.commit()
                    result = cursor.fetchone()
                    wyb_all_num = str(result[0])

                    #  wyb_num     ：NIO 编号
                    #  wyb_all_num ：内码
                    #  wyb_style   ：万用表型号



                    sql = "SELECT dang_num FROM wyb_info WHERE type = %s"
                    cursor.execute(sql, (wyb_style,))
                    db.commit()
                    result = cursor.fetchone()
                    dang_num = int(result[0])

                    # 清空原来位置的万用表数据

                    sql2 = "INSERT INTO new2( num, nio, equip_num, dang_num )SELECT num, nio, equip_num, dang_num FROM wyb_test_info WHERE equip_num > %s"
                    sql3 = "UPDATE wyb_test_info SET dang_num = NULL,nio = NULL,equip_num = NULL,num = NULL WHERE equip_num > %s"


                    k_pos = end_pos - 1
                    cursor.execute(sql2,(end_pos,))
                    db.commit()
                    cursor.execute(sql3, (k_pos,))
                    db.commit()


                    # 插入新的万用表
                    sql = "SELECT max(id) FROM wyb_test_info WHERE nio IS NOT NULL"
                    cursor.execute(sql)
                    db.commit()
                    result = cursor.fetchone()
                    max_id = result[0]

                    if max_id is None:

                        max_id = 1

                    max_id = int(max_id)

                    for (i, j) in zip(range(max_id + 1, dang_num + max_id + 1), range(1, dang_num + 1)):

                        sql1 = "UPDATE wyb_test_info SET num = %s WHERE id = %s"
                        sql2 = "UPDATE wyb_test_info SET nio = %s WHERE id = %s"
                        sql3 = "UPDATE wyb_test_info SET dang_num = %s WHERE id = %s"
                        sql4 = "UPDATE wyb_test_info SET equip_num = %s WHERE id = %s"

                        cursor.execute(sql1, (wyb_all_num, str(i)))
                        db.commit()
                        cursor.execute(sql2, (wyb_num, str(i)))
                        db.commit()
                        cursor.execute(sql3, (str(j), str(i)))
                        db.commit()
                        cursor.execute(sql4, (str(end_pos), str(i)))
                        db.commit()

                    # 还原数据

                    sql4 = "DELETE FROM wyb_test_info WHERE equip_num IS NULL"
                    sql5 = "INSERT INTO wyb_test_info (num, nio, equip_num, dang_num) SELECT num, nio, equip_num, dang_num FROM new2"
                    sql6 = "SET @auto_id = 0"
                    sql7 = "UPDATE wyb_test_info SET id = (@auto_id := @auto_id + 1)"
                    sql8 = "ALTER TABLE wyb_test_info AUTO_INCREMENT = 1"
                    sql9 = "INSERT INTO wyb_test_info(num) VALUES(NULL)"
                    sql10 = "DELETE FROM new2"

                    cursor.execute(sql4)
                    db.commit()
                    cursor.execute(sql5)
                    db.commit()
                    cursor.execute(sql6)
                    db.commit()
                    cursor.execute(sql7)
                    db.commit()
                    cursor.execute(sql8)
                    db.commit()
                    for i in range(1,30):
                        cursor.execute(sql9)
                        db.commit()
                    cursor.execute(sql10)
                    db.commit()


                    vector_wyb[end_pos - 1] = [end_pos, wyb_num, wyb_style]


                # 万用表提示
                message = "扫码完成，已将 {} 该编号万用表 配对于 {} 号 装夹位，请开始人工装夹".format(wyb_num, wyb_style)


                newItem = QTableWidgetItem(wyb_num)  # 创建表格项---文本项目
                self.table.setItem(end_pos, 1, newItem)  # 给指定单元格设置数据

                newItem1 = QTableWidgetItem(wyb_style)  # 创建表格项---文本项目
                self.table.setItem(end_pos, 2, newItem1)  # 给指定单元格设置数据


                QMessageBox.information(QMessageBox(), '装夹', message)
                QMessageBox.information(QMessageBox(), '装夹', '安装完毕！')

            else:

                QMessageBox.information(None, '提示', '未输入装夹位，请重新点击 修改装夹！')

        else:

            return


    def alter_image(self):

        self.pixmap = QPixmap(equip_image)
        self.new_pixmap = self.pixmap.scaled(self.equip_graph.width(), self.equip_graph.height())
        self.equip_graph.setPixmap(self.new_pixmap)


    def equip_exit(self):


        self.zhujiemian = wyb()
        self.zhujiemian.ui.show()
        self.ui.close()







class find_result(QWidget):


    def __init__(self):

        super(find_result, self).__init__()

        self.ui = QUiLoader().load(addr_ui + "/find_result.ui")

        self.ui.find_result.clicked.connect(self.find_result)  # 下一步

        self.ui.find_download.clicked.connect(self.find_download)

        self.ui.find_exit.clicked.connect(self.find_exit)

        for i in range(15):

            for j in range(25):
                self.ui.find_result_table.setItem(j + 1, i, None)



        if find_wyb_way == 'nio':

            self.ui.find_input_type.setVisible(False)
            self.ui.find_year.setVisible(False)
            self.ui.find_month.setVisible(False)
            self.ui.find_day.setVisible(False)
            self.ui.label1.setVisible(False)
            self.ui.label2.setVisible(False)
            self.ui.label3.setVisible(False)
            self.ui.find_label2.setVisible(False)
            self.ui.find_label3.setVisible(False)


        elif find_wyb_way == 'type':

            self.ui.find_input_nio.setVisible(False)
            self.ui.find_year.setVisible(False)
            self.ui.find_month.setVisible(False)
            self.ui.find_day.setVisible(False)
            self.ui.label1.setVisible(False)
            self.ui.label2.setVisible(False)
            self.ui.label3.setVisible(False)
            self.ui.find_label1.setVisible(False)
            self.ui.find_label3.setVisible(False)


        elif find_wyb_way == 'type_day':

            self.ui.find_input_nio.setVisible(False)
            self.ui.find_label1.setVisible(False)

            self.ui.find_label2.move(70, 110)
            self.ui.find_input_type.move(170, 110)

            self.ui.find_label3.move(460, 110)
            self.ui.find_year.move(560, 120)
            self.ui.label1.move(640, 120)
            self.ui.find_month.move(670, 120)
            self.ui.label3.move(730, 120)
            self.ui.find_day.move(750, 120)
            self.ui.label2.move(810, 120)


        elif find_wyb_way == 'day':

            self.ui.find_input_nio.setVisible(False)
            self.ui.find_input_type.setVisible(False)
            self.ui.find_label1.setVisible(False)
            self.ui.find_label2.setVisible(False)

        linshi_addr = "QLabel{border-image: url(" + addr_ui + "/登录底板.png)}"
        self.ui.find_result_pic.setStyleSheet(linshi_addr)


    def find_result(self):

        global find_result_name

        nio = self.ui.find_input_nio.text()
        type = self.ui.find_input_type.text()
        year = self.ui.find_year.currentText()
        month = f"{(int(self.ui.find_month.currentText())):02}"
        day = f"{(int(self.ui.find_day.currentText())):02}"

        wyb_time = year + month + day

        print(wyb_time)

        for i in range(15):

            for j in range(25):
                self.ui.find_result_table.setItem(j + 1, i, None)



        cursor = db.cursor()

        if wyb_time != '20240101':

            if (type != "") and (nio != ""):

                QMessageBox.information(QMessageBox(), '查询', "输入错误！ nio和型号只能选择其一，不可都输入！")

            else:

                if (type == "") and (nio == ""):

                    sql = "SELECT num, mea_time, unit, measure, m_1, m_2, m_3, m_average, m_max, " \
                          "m_min, m_sz, m_xd, m_repeat FROM wyb_value " \
                          "WHERE mea_time IS NOT NULL " \
                          "AND LEFT(mea_time, 8) = %s"

                    # 执行查询，将wyb_num和a1作为参数传递
                    cursor.execute(sql, (wyb_time,))
                    db.commit()
                    wyb_measure = cursor.fetchall()

                    length1 = len(wyb_measure)

                    for i in range(0, length1):

                        self.ui.find_result_table.setItem(i + 1, 1, QTableWidgetItem(wyb_measure[i][0]))

                        sql = "SELECT nio FROM wyb_name_info WHERE num = %s"
                        cursor.execute(sql, wyb_measure[i][0])
                        db.commit()
                        result = cursor.fetchone()
                        wyb_nio = str(result[0])

                        self.ui.find_result_table.setItem(i + 1, 0, QTableWidgetItem(wyb_nio))

                        for j in range(0, 12):

                            self.ui.find_result_table.setItem(i + 1, j + 2, QTableWidgetItem(str(wyb_measure[i][j +1 ])))

                    find_result_name = wyb_time

                elif type == "":

                    QMessageBox.information(QMessageBox(), '查询', "输入错误！ 请输入型号！")

                elif nio == "":

                    sql = "SELECT nio,num FROM wyb_name_info WHERE type = %s"
                    cursor.execute(sql, type)
                    db.commit()
                    wyb_all_mea = cursor.fetchall()

                    len_key = 0

                    for mea_info in wyb_all_mea:

                        nio, wyb_num = mea_info  # 解包元组获取nio和wyb_num

                        # 使用获取的wyb_num查询wyb_value表
                        sql = "SELECT mea_time, unit, measure, m_1, m_2, m_3, m_average, m_max, " \
                              "m_min, m_sz, m_xd, m_repeat FROM wyb_value " \
                              "WHERE num = %s AND mea_time IS NOT NULL " \
                              "AND LEFT(mea_time, 8) = %s"

                        # 执行查询，将wyb_num和a1作为参数传递
                        cursor.execute(sql, (wyb_num, wyb_time))
                        db.commit()
                        wyb_measure = cursor.fetchall()

                        # 假设wyb_measure不为空
                        if wyb_measure:
                            length2 = len(wyb_measure)

                            # 遍历wyb_measure列表并填充表格
                            for i in range(length2):
                                # 这里假设表格的列数足够，且nio和wyb_num不变化
                                self.ui.find_result_table.setItem(i + len_key + 1, 0, QTableWidgetItem(str(nio)))
                                self.ui.find_result_table.setItem(i + len_key + 1, 1, QTableWidgetItem(str(wyb_num)))

                                for j in range(12):
                                    # 这里假设wyb_measure[i]是一个包含12个元素的列表或元组
                                    self.ui.find_result_table.setItem(i + len_key + 1, j + 2,
                                                                      QTableWidgetItem(str(wyb_measure[i][j])))

                            len_key = length2


                    find_result_name = type + '_' + wyb_time


        elif wyb_time == '20240101':

            if (type != "") and (nio != ""):

                QMessageBox.information(QMessageBox(), '查询', "输入错误！ nio和型号只能选择其一，不可都输入！")

            else:

                if (type == "") and (nio == ""):

                    QMessageBox.information(QMessageBox(), '查询', "输入错误！ 请输入nio或型号！")

                elif type == "":

                    sql = "SELECT num,type FROM wyb_name_info WHERE nio = %s"
                    cursor.execute(sql, nio)
                    db.commit()
                    result = cursor.fetchone()
                    wyb_num = str(result[0])
                    wyb_type = str(result[1])

                    sql = "SELECT mea_time, unit, measure, m_1, m_2, m_3, m_average, m_max, " \
                          "m_min, m_sz, m_xd, m_repeat FROM wyb_value WHERE num = %s AND mea_time IS NOT NULL"
                    cursor.execute(sql, wyb_num)
                    db.commit()
                    wyb_measure = cursor.fetchall()
                    length1 = len(wyb_measure)


                    for i in range(0, length1):

                        self.ui.find_result_table.setItem(i + 1, 0, QTableWidgetItem(str(nio)))
                        self.ui.find_result_table.setItem(i + 1, 1, QTableWidgetItem(str(wyb_num)))

                        for j in range(0, 12):

                            self.ui.find_result_table.setItem(i + 1, j + 2, QTableWidgetItem(str(wyb_measure[i][j])))


                    find_result_name = nio


                elif nio == "":

                    sql = "SELECT nio,num FROM wyb_name_info WHERE type = %s"
                    cursor.execute(sql, type)
                    db.commit()
                    wyb_all_mea = cursor.fetchall()

                    len_key = 0

                    for mea_info in wyb_all_mea:

                        nio, wyb_num = mea_info  # 解包元组获取nio和wyb_num

                        # 使用获取的wyb_num查询wyb_value表
                        sql = "SELECT mea_time, unit, measure, m_1, m_2, m_3, m_average, m_max, " \
                              "m_min, m_sz, m_xd, m_repeat FROM wyb_value WHERE num = %s AND mea_time IS NOT NULL"

                        cursor.execute(sql, wyb_num)
                        db.commit()
                        wyb_measure = cursor.fetchall()

                        # 假设wyb_measure不为空
                        if wyb_measure:
                            length2 = len(wyb_measure)

                            # 遍历wyb_measure列表并填充表格
                            for i in range(length2):
                                # 这里假设表格的列数足够，且nio和wyb_num不变化
                                self.ui.find_result_table.setItem(i + len_key + 1, 0, QTableWidgetItem(str(nio)))
                                self.ui.find_result_table.setItem(i + len_key + 1, 1, QTableWidgetItem(str(wyb_num)))

                                for j in range(12):
                                    # 这里假设wyb_measure[i]是一个包含12个元素的列表或元组
                                    self.ui.find_result_table.setItem(i + len_key + 1, j + 2, QTableWidgetItem(str(wyb_measure[i][j])))

                            len_key = length2

                    find_result_name = type



    def find_download(self):

        folder_path = QFileDialog.getExistingDirectory(self, "选择导出路径", "", QFileDialog.Option.ShowDirsOnly)

        if folder_path:  # 确保用户选择了文件夹

            wb = openpyxl.Workbook()
            ws = wb.active

            # 写入表头
            for i in range(self.ui.find_result_table.columnCount()):
                header_item = self.ui.find_result_table.horizontalHeaderItem(i)
                ws.cell(row=1, column=i + 1, value=header_item.text())  # 表头从第1行开始

            # 写入数据
            for row in range(self.ui.find_result_table.rowCount()):
                # 由于表头已经占用了第1行，数据从第2行开始写入
                for column in range(self.ui.find_result_table.columnCount()):
                    cell_item = self.ui.find_result_table.item(row, column)
                    ws.cell(row=row + 1, column=column + 1, value=cell_item.text() if cell_item else "")

            for col in range(1, ws.max_column + 1):  # ws.max_column是列数的最大值
                ws.column_dimensions[get_column_letter(col)].auto_size = True

            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20

            # 设置行高以自适应内容
            for row in range(1, ws.max_row + 1):  # ws.max_row是行数的最大值
                ws.row_dimensions[row].auto_size = True

            # 创建文件名，格式为 output_加上时间戳
            file_name = f"output_{find_result_name}"

            # 完整的文件路径
            file_path = folder_path + '/' + file_name + ".xlsx"


            wb.save(file_path)

            QMessageBox.information(QMessageBox(), '查询', "导出成功！")




    def find_exit(self):


        self.zhujiemian = find_way()
        self.zhujiemian.ui.show()
        self.ui.close()






class run():


    def __init__(self):

        self.ui = QUiLoader().load(addr_ui + "/run.ui")

        self.ui.run_exit.clicked.connect(self.run_exit)  # 下一步

        self.ui.run_stop.clicked.connect(self.run_stop)  # 下一步

        self.ui.run_start.clicked.connect(self.run_start_system)

        self.key = 0

        self.run_table = self.ui.run_table

        self.run_clear_wyb()


        self.run_show_wyb_info(1)




    # 前期验证 + 显示信息


    def run_show_wyb_info(self, wyb_work_num):

        # 一、更新屏幕万用表信息
        wyb_info_num = '第 ' + str(wyb_work_num) + ' 台'
        self.ui.run_wyb_num.setText(wyb_info_num)

        wyb_work_num = str(wyb_work_num)

        cursor = db.cursor()
        sql = "SELECT num,nio FROM wyb_test_info WHERE equip_num = %s"
        cursor.execute(sql, wyb_work_num)
        db.commit()
        result = cursor.fetchone()
        wyb_type = str(result[0])
        wyb_nio = str(result[1])

        self.ui.run_wyb_type.setText(wyb_type)
        self.ui.run_wyb_nio.setText(wyb_nio)

        # 二、万用表数据库

        sql = "SELECT unit FROM wyb_value WHERE num = %s"
        cursor.execute(sql, wyb_type)
        db.commit()
        wyb_unit = cursor.fetchall()

        sql = "SELECT measure FROM wyb_value WHERE num = %s"
        cursor.execute(sql, wyb_type)
        db.commit()
        wyb_measure = cursor.fetchall()

        length1 = len(wyb_measure)

        for i in range(0, length1):
            self.ui.run_table.setItem(0, i + 1, QTableWidgetItem(str(wyb_unit[i][0])))
            self.ui.run_table.setItem(1, i + 1, QTableWidgetItem(str(wyb_measure[i][0])))


    def run_modify_work_start_wyb_num(self):

        # 1、拍当前装夹位，计算得到转盘转动格数，使 工作位 转至 1 号 位

        photo_name = "robot_equip_pic"
        testGrabImage.paizhao(photo_name)
        time.sleep(2)
        dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg")
        equip_num = int(dedaozhi)

        if equip_num in [1, 2, 3]:

            robot_act_new.send_zp_start(4 - equip_num)

        elif equip_num in [5, 6, 7, 8]:

            robot_act_new.send_zp_start(12 - equip_num)

        # 2、验证 工作位 为 1 号 位（装夹位 为 4 号 位）

        time.sleep(3)
        photo_name_again = "robot_equip_pic_again"
        testGrabImage.paizhao(photo_name_again)
        time.sleep(2)
        dedaozhi_again = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name_again + ".jpg")
        equip_num_again = int(dedaozhi_again)
        print(equip_num_again)

        return equip_num_again


    def run_find_kong_zp_pos(self,type,wyb_num):

        global tool_x, tool_y

        cursor = db.cursor()

        cursor.execute("SELECT d_all_wyb FROM jiajv_info WHERE id = 1")
        db.commit()
        result = cursor.fetchone()
        x1, y1, z1, rx1, ry1, rz1 = map(float, result[0].split('&'))
        W_photo = [x1, y1, z1, rx1, ry1, rz1]

        robot_act_new.movej_PTP_getJ(3, 50.0, W_photo)

        time.sleep(5)

        testGrabImage.paizhao(wyb_num,0)

        convert_new.pic_equip_path = "E:/photo_test/" + wyb_num + ".jpg"
        convert_new.pic_blue_path = "E:/photo_test/" + type + "_ori.jpg"

        time.sleep(1)

        #0.241935

        kong_loc = convert_new.get_now_wyb_loc(4, 0.243)[1]
        zp_loc = convert_new.get_now_wyb_loc(4, 0.243)[3]


        k1_loc_x = x1 - kong_loc[0][1] + tool_x
        k1_loc_y = kong_loc[0][0] + y1 + tool_y
        k2_loc_x = x1 - kong_loc[1][1] + tool_x
        k2_loc_y = kong_loc[1][0] + y1 + tool_y
        k3_loc_x = x1 - kong_loc[2][1] + tool_x
        k3_loc_y = kong_loc[2][0] + y1 + tool_y
        k4_loc_x = x1 - kong_loc[3][1] + tool_x
        k4_loc_y = kong_loc[3][0] + y1 + tool_y

        zp_loc_x = x1 - zp_loc[1] + tool_x + 1.7
        zp_loc_y = zp_loc[0] + y1 + tool_y - 3.5

        print('y1')
        print(y1)
        print('kong_loc')
        print(kong_loc[0][0])



        k1_loc_x = round(k1_loc_x, 3)
        k1_loc_y = round(k1_loc_y, 3)
        k2_loc_x = round(k2_loc_x, 3)
        k2_loc_y = round(k2_loc_y, 3)
        k3_loc_x = round(k3_loc_x, 3)
        k3_loc_y = round(k3_loc_y, 3)
        k4_loc_x = round(k4_loc_x, 3)
        k4_loc_y = round(k4_loc_y, 3)

        zp_loc_x = round(zp_loc_x, 3)
        zp_loc_y = round(zp_loc_y, 3)

        k1_pos = f"{k1_loc_x}&{k1_loc_y}"
        k2_pos = f"{k2_loc_x}&{k2_loc_y}"
        k3_pos = f"{k3_loc_x}&{k3_loc_y}"
        k4_pos = f"{k4_loc_x}&{k4_loc_y}"
        zp_pos = f"{zp_loc_x}&{zp_loc_y}"


        sql = "UPDATE wyb_info SET k2_pos = %s WHERE type = %s"
        cursor.execute(sql, (k1_pos, type))
        db.commit()
        sql = "UPDATE wyb_info SET k4_pos = %s WHERE type = %s"
        cursor.execute(sql, (k2_pos, type))
        db.commit()
        sql = "UPDATE wyb_info SET k1_pos = %s WHERE type = %s"
        cursor.execute(sql, (k3_pos, type))
        db.commit()
        sql = "UPDATE wyb_info SET k3_pos = %s WHERE type = %s"
        cursor.execute(sql, (k4_pos, type))
        db.commit()

        '''
        sql = "UPDATE wyb_info SET zp_pos = %s WHERE type = %s"
        cursor.execute(sql, (zp_pos, type))
        db.commit()
        '''

        print(zp_pos)


    def run_modify_work_next_wyb_num(self, modify_erm_num):

        self.P_modify_height = copy.copy(robot_act_new.P_modify)

        self.P_modify_height[2] = -80.000

        robot_act_new.movej_PTP_getJ(3, 40.0, self.P_modify_height)

        robot_act_new.movej_PTP_getJ(3, 70.0, robot_act_new.P_modify)

        time.sleep(2)

        # 扫码出图
        photo_name = "modify_work_ewm"
        testGrabImage.paizhao(photo_name)
        time.sleep(2)
        dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg")
        equip_num = int(dedaozhi)
        print('得到值：')
        print(equip_num)
        if equip_num != modify_erm_num:

            return False

        elif equip_num == modify_erm_num:

            return True








    # 开始运行


    def run_start_system(self):


        QMessageBox.information(QMessageBox(), '运行', '请确认将 双门 关闭！机器人开始运行！')


        # 复位 和 验证

        #self.equip_num_again = self.run_modify_work_start_wyb_num()


        self.equip_num_again = 4

        print('当前工作位：')
        print(self.equip_num_again)



        if self.equip_num_again != 4:

            QMessageBox.information(QMessageBox(), '运行', '目前工位不为 1号工位，请检查转盘！')

        elif self.equip_num_again == 4:

            self.run_clear_wyb()

            self.run_show_wyb_info(1)

            wyb_system_start = threading.Thread(target=self.run_start_wyb)

            wyb_system_start.start()



    def run_start_wyb(self):


        cursor = db.cursor()

        # 查 最大装夹数
        cursor.execute('SELECT max(equip_num) FROM wyb_test_info')
        db.commit()
        max_equip_num = int(cursor.fetchone()[0])

        for equip_num in range(1, max_equip_num + 1):

            print(equip_num)

            # 查 内码 num
            cursor.execute('SELECT num FROM wyb_test_info WHERE equip_num = %s',(equip_num,))
            db.commit()
            wyb_num = cursor.fetchone()[0]

            # 查 型号 type
            cursor.execute('SELECT type FROM wyb_name_info WHERE num = %s', (wyb_num,))
            db.commit()
            wyb_type = cursor.fetchone()[0]


            # 拍照-标定-4孔+拨盘中心

            #self.run_find_kong_zp_pos(wyb_type, wyb_num)


            # 插针

            print('插孔')
            #robot_act_new.jiajv(1, wyb_type, 0)
            #robot_act_new.jiajv(2, wyb_type, 0)
            #robot_act_new.jiajv(3, wyb_type, 0)
            #robot_act_new.jiajv(4, wyb_type, 0)


            time.sleep(2)



            # 关闭标准源通道
            new_bzy.close_door()


            # 测量 第一台 万用表 FLUKE
            self.run_mea_vc_test()

            print('测量')






            # 还针
            #robot_act_new.jiajv_back(4)
            #robot_act_new.jiajv_back(3)
            #robot_act_new.jiajv_back(2)
            #robot_act_new.jiajv_back(1)

            time.sleep(2)
    



            if equip_num != max_equip_num :
                
                
                # 验证转盘
    
                robot_act_new.send_zp_start(1)

                if equip_num >= 1 and equip_num <= 6:

                    modify_key = self.run_modify_work_next_wyb_num(equip_num + 2)

                else:

                    modify_key = self.run_modify_work_next_wyb_num(equip_num - 6)


        
                if modify_key is False:

                    QMessageBox.information(QMessageBox(), '测量', '当前转盘出现问题！')

                    return

                elif modify_key is True:

                    #QMessageBox.information(QMessageBox(), '测量', '2号万用表开始测量！')

                    self.run_clear_wyb()

                    self.run_show_wyb_info(equip_num + 1)

                    print('验证成功:equip')












    # 万用表不同型号


    def run_mea_fluke(self):

        # 测 ！！！

        # 转拨盘，此时移到了lcd位置

        begin_angle,end_angle = self.run_find_dang_angle(0, 2, 'FLUKE&17B+')


        robot_act_new.execute_robot_rotate(begin_angle, end_angle, 'FLUKE&17B+')


        # 测 v
        show_unit = 'DCV+'
        mea_unit = 'V'
        mea_value = 4

        self.run_mea_same(show_unit, mea_unit, mea_value, vector_v1, bzy_v1, 5, '4_3')

        # 记录测量单位
        vector_v1[3] = show_unit
        vector_v1[4] = mea_value

        # 机械臂旋转下一档
        begin_angle, end_angle = self.run_find_dang_angle(2, 3, 'FLUKE&17B+')

        robot_act_new.execute_robot_rotate(begin_angle, end_angle, 'FLUKE&17B+')

        # 测 mv
        show_unit = 'MDCV+'
        mea_unit = 'mV'
        mea_value = 400

        self.run_mea_same(show_unit, mea_unit, mea_value, vector_v2, bzy_v2, 13, '4_1')

        vector_v2[3] = show_unit
        vector_v2[4] = mea_value

        begin_angle, end_angle = self.run_find_dang_angle(3, 6, 'FLUKE&17B+')

        robot_act_new.execute_robot_rotate(begin_angle, end_angle, 'FLUKE&17B+')

        # 测 ma
        show_unit = 'DCA+'
        mea_unit = 'A'
        mea_value = 3

        self.run_mea_same(show_unit, mea_unit, mea_value, vector_a1, bzy_a1, 22, '4_3')

        vector_a1[3] = show_unit
        vector_a1[4] = mea_value + 1

        # robot_act_new.execute_robot_rotate(1, 135.000, 157.500, '17B+', 1)

        show_unit = 'MDCA+'
        mea_unit = 'mA'
        mea_value = 40

        self.run_mea_same(show_unit, mea_unit, mea_value, vector_a2, bzy_a2, 28, '4_3')

        vector_a2[3] = show_unit
        vector_a2[4] = mea_value

        show_unit = 'MDCA+'
        mea_unit = 'mA'
        mea_value = 400

        self.run_mea_same(show_unit, mea_unit, mea_value, vector_a3, bzy_a3, 29, '4_3')

        vector_a3[3] = show_unit
        vector_a3[4] = mea_value



        # 计 算 ！！！

        self.caculator_num(bzy_v1, vector_v1, 5)
        self.caculator_num(bzy_v2, vector_v2, 13)
        self.caculator_num(bzy_a1, vector_a1, 22)
        self.caculator_num(bzy_a2, vector_a2, 28)
        self.caculator_num(bzy_a3, vector_a3, 29)


        # 转盘回位！！！
        begin_angle, end_angle = self.run_find_dang_angle(6, 0, 'FLUKE&17B+')

        robot_act_new.execute_robot_rotate(begin_angle, end_angle, 'FLUKE&17B+')




        # 下一台！！！
        time.sleep(6)



    def run_mea_vc(self,all_mea_num):


        global k_num,be,en


        cursor = db.cursor()

        begin_angle, end_angle = self.run_find_dang_angle(0, 1, 'VICTOR&VC890C+')

        robot_act_new.execute_robot_rotate(begin_angle, end_angle, 'VICTOR&VC890C+')

        for i in range(1, all_mea_num + 1):

            sql = "SELECT dang_num FROM wyb_value WHERE num = %s and id_num = %s"
            cursor.execute(sql, ('VICTORVC890C+001', i))
            db.commit()
            result = cursor.fetchone()

            dang_num = int(result[0])

            if dang_num != k_num:

                k_num = dang_num

                be = be + 1
                en = en + 1

                begin_angle, end_angle = self.run_find_dang_angle(be, en, 'VICTOR&VC890C+')

                robot_act_new.execute_robot_rotate(begin_angle, end_angle, 'VICTOR&VC890C+')


            sql = "SELECT unit,measure,hz FROM wyb_value WHERE id_num = %s and num = %s"

            cursor.execute(sql, (i,'VICTORVC890C+001'))
            db.commit()
            result = cursor.fetchone()
            show_unit = str(result[0])
            mea_value = str(result[1])
            hz_value = int(result[2])

            mea_unit = self.run_check_num(show_unit)


            print(mea_value)

            if hz_value != 0:

                self.run_mea_same(show_unit, mea_unit, mea_value, vector_v1, bzy_v1, i, hz_value)

            else:

                self.run_mea_same(show_unit, mea_unit, mea_value, vector_v1, bzy_v1, i, 0)

            # 记录测量单位
            vector_v1[3] = show_unit
            vector_v1[4] = mea_value

            self.caculator_num(bzy_v1, vector_v1, i)



        # 转盘回位！！！
        begin_angle, end_angle = self.run_find_dang_angle(en, 0, 'VICTOR&VC890C+')

        robot_act_new.execute_robot_rotate(begin_angle, end_angle - 1, 'VICTOR&VC890C+')




        # 下一台！！！
        time.sleep(6)




    def run_mea_vc_test(self):


        cursor = db.cursor()

        sql = "SELECT unit,measure,hz FROM wyb_value WHERE id_num = %s and num = %s"

        cursor.execute(sql, (21, 'VICTORVC890C+001'))
        db.commit()
        result = cursor.fetchone()
        show_unit = str(result[0])
        mea_value = str(result[1])
        hz_value = int(result[2])

        mea_unit = self.run_check_num(show_unit)

        print(mea_value)
        print(1)
        print(mea_unit)
        if hz_value != 0:

            self.run_mea_same(show_unit, mea_unit, '200', vector_v1, bzy_v1, 21, hz_value)

        else:

            self.run_mea_same(show_unit, mea_unit, '200', vector_v1, bzy_v1, 21, 0)

        # 记录测量单位
        vector_v1[3] = show_unit
        vector_v1[4] = mea_value

        self.caculator_num(bzy_v1, vector_v1, 21)

    # 辅助函数


    def run_check_num(self, a):

        if 'M' in a and 'V' in a:
            return 'mV'
            # 检查字符串a是否含有'V'
        elif 'V' in a:
            return 'V'
            # 检查字符串a是否含有'A'且不含有'V'
        elif 'M' in a and 'A' in a:
            return 'mA'
            # 检查字符串a是否都含有'U'和'A'
        elif 'U' in a and 'A' in a:
            return 'uA'
            # 检查字符串a是否含有'R'和'K'
        elif 'A' in a and 'V' not in a:
            return 'A'
            # 检查字符串a是否都含有'M'和'A'
        elif 'R' in a and 'K' in a:
            return 'KOHM'
            # 检查字符串a是否含有'R'和'M'
        elif 'R' in a and 'M' in a:
            return 'MOHM'
            # 检查字符串a是否含有'R'且不含有'M'和'K'
        elif 'R' in a and 'M' not in a and 'K' not in a:
            return 'OHM'



    def run_find_dang_angle(self,begin,end,num):

        cursor = db.cursor()
        sql = "SELECT dang_angle FROM wyb_unit_info WHERE dang_num = %s and type = %s"

        cursor.execute(sql, (begin,num))
        db.commit()
        result = cursor.fetchone()
        begin_angle = float(result[0])

        cursor.execute(sql, (end,num))
        db.commit()
        result = cursor.fetchone()
        end_angle = float(result[0])

        return begin_angle,end_angle








    # 连续测量三次


    def run_mea_same(self,show_unit,mea_unit,mea_value,vector_v,bzy_v,table_lie,hz):



        self.ui.run_wyb_label_unit.setText(show_unit)
        self.ui.run_wyb_label_value.setText(str(mea_value))
        self.ui.run_wyb_label_result.setText(' ')
        # 测量值，测量单位，表格第几行，表格第几列，几位数保留小数位几位
        vector_v[0], bzy_v[0] = self.run_mea_unit(mea_value, mea_unit, 2, table_lie, hz)
        self.ui.run_wyb_label_result.setText(str(vector_v[0]))
        # 写入数据库
        cursor = db.cursor()
        sql1 = "UPDATE wyb_value SET m_1 = %s WHERE unit = %s AND measure = %s"
        cursor.execute(sql1, (str(vector_v[0]), show_unit, str(mea_value)))
        db.commit()
        # 写入时间
        sql1 = "UPDATE wyb_value SET mea_time = %s WHERE unit = %s AND measure = %s"
        cursor.execute(sql1, ((datetime.now()).strftime("%Y%m%d%H%M%S"), show_unit, str(mea_value)))
        db.commit()

        time.sleep(2)

        self.ui.run_wyb_label_unit.setText(show_unit)
        self.ui.run_wyb_label_value.setText(str(mea_value))
        self.ui.run_wyb_label_result.setText(' ')
        # 测量值，测量单位，表格第几行，表格第几列，几位数保留小数位几位
        vector_v[1], bzy_v[1] = self.run_mea_unit(mea_value, mea_unit, 3, table_lie, hz)
        self.ui.run_wyb_label_result.setText(str(vector_v[1]))
        # 写入数据库
        sql1 = "UPDATE wyb_value SET m_2 = %s WHERE unit = %s AND measure = %s"
        cursor.execute(sql1, (str(vector_v[1]), show_unit, str(mea_value)))
        db.commit()

        time.sleep(2)

        self.ui.run_wyb_label_unit.setText(show_unit)
        self.ui.run_wyb_label_value.setText(str(mea_value))
        self.ui.run_wyb_label_result.setText(' ')
        # 测量值，测量单位，表格第几行，表格第几列，几位数保留小数位几位
        vector_v[2], bzy_v[2] = self.run_mea_unit(mea_value, mea_unit, 4, table_lie, hz)
        self.ui.run_wyb_label_result.setText(str(vector_v[2]))
        # 写入数据库
        sql1 = "UPDATE wyb_value SET m_3 = %s WHERE unit = %s AND measure = %s"
        cursor.execute(sql1, (str(vector_v[2]), show_unit, str(mea_value)))
        db.commit()

        time.sleep(2)




    def run_mea_unit(self, num, mea, hang, lie, hz):



        '''

                if num == '200' and mea == 'mA':

                    if hz == 0:
                        new_bzy.mea_dc('20mA')
                        time.sleep(5)
                        new_bzy.close_door()
                        time.sleep(3)




                # 给标准源输出 30V 电压，更新表格
                if hz == 0:
                    mea_value =  str(num) + mea
                    print('测量单位')
                    print(mea_value)
                    bzy_mea = new_bzy.mea_dc(mea_value)
                elif hz != 0:
                    mea_value = str(num) + mea
                    print('hz单位')
                    print(mea_value)
                    bzy_mea = new_bzy.mea_ac(mea_value, hz)




                bzy_mea = float(bzy_mea)
                time.sleep(4)

                # 识别 lcd 数字
                photo_name = str(num) + mea + '_' + datetime.now().strftime("%Y%m%d%H%M%S")
                testGrabImage.paizhao(photo_name)
                time.sleep(1)
                self.run_image = "E:/photo/" + photo_name + ".jpg"



                mea_30 = ewm_and_lcd.lcd_shibie("E:/photo/" + photo_name + ".jpg")[0]

                delta_bzy_mea = 0

                if 'L' in key_pd:

                    new_bzy.close_door()

                    newItem = QTableWidgetItem('超量程')
                    self.ui.run_table.setItem(11, lie, newItem)

                    if '-' in num:

                        len_value = len(num) - 1
                        num = int(num)
                        if len_value == 3:
                            num = num + 1
                            delta_bzy_mea = -1
                        elif len_value == 2:
                            num = num + 0.1
                            delta_bzy_mea = -0.1
                        elif len_value == 1:
                            num = num + 0.01
                            delta_bzy_mea = -0.01
                        elif len_value == 4:
                            num = num + 10
                            delta_bzy_mea = -10
                    else:
                        len_value = len(num)
                        num = int(num)
                        if len_value == 3:
                            num = num - 1
                            delta_bzy_mea = 1
                        elif len_value == 2:
                            num = num - 0.1
                            delta_bzy_mea = 0.1
                        elif len_value == 1:
                            num = num - 0.01
                            delta_bzy_mea = 0.01
                        elif len_value == 4:
                            num = num - 10
                            delta_bzy_mea = 10

                    if hz == 0:
                        mea_value = str(num) + mea
                        print('测量单位')
                        print(mea_value)
                        bzy_mea = new_bzy.mea_dc(mea_value)
                    elif hz != 0:
                        mea_value = str(num) + mea
                        print('hz单位')
                        print(mea_value)
                        bzy_mea = new_bzy.mea_ac(mea_value, hz)

                    bzy_mea = float(bzy_mea)
                    time.sleep(4)

                    # 识别 lcd 数字
                    photo_name = str(num) + mea + '_' + datetime.now().strftime("%Y%m%d%H%M%S")
                    testGrabImage.paizhao(photo_name)
                    time.sleep(1)
                    self.run_image = "E:/photo/" + photo_name + ".jpg"

                    mea_30 = ewm_and_lcd.lcd_shibie("E:/photo/" + photo_name + ".jpg")[0]




                # 处理 lcd 读数，加小数点，并显示于 ui 界面上
                mea_30 = int(mea_30)
                mea_30 = float(mea_30)


                xunhuan = 0

                while (True):

                    print('bzy_mea')
                    print(bzy_mea)
                    value_test = mea_30 / bzy_mea

                    if (value_test > 0.8 and value_test < 1.2) or (value_test < -0.8 and value_test > -1.2):

                        break

                    else:

                        mea_30 = mea_30 * 0.1

                        xunhuan = xunhuan + 1

                        if xunhuan > 5:

                            print('sbsb')
                            mea_30 = mea_30 * 100000
                            break


                str_mea_num = len(str(int(bzy_mea)))

                if '-' in str(int(bzy_mea)):

                    str_mea_num = str_mea_num - 1
                    mea_30 = - mea_30


                print('mea_qian')
                print(mea_30)



                if delta_bzy_mea == -0.1:
                    mea_30 = mea_30 - 0.1
                    bzy_mea = bzy_mea - 0.1
                elif delta_bzy_mea == -1:
                    mea_30 = mea_30 - 1
                    bzy_mea = bzy_mea - 1
                elif delta_bzy_mea == -0.01:
                    mea_30 = mea_30 - 0.01
                    bzy_mea = bzy_mea - 0.01
                elif delta_bzy_mea == -10:
                    mea_30 = mea_30 - 10
                    bzy_mea = bzy_mea - 10
                elif delta_bzy_mea == 0.1:
                    mea_30 = mea_30 + 0.1
                    bzy_mea = bzy_mea + 0.1
                elif delta_bzy_mea == 1:
                    mea_30 = mea_30 + 1
                    bzy_mea = bzy_mea + 1
                elif delta_bzy_mea == 10:
                    mea_30 = mea_30 + 10
                    bzy_mea = bzy_mea + 10
                elif delta_bzy_mea == 0.01:
                    mea_30 = mea_30 + 0.01
                    bzy_mea = bzy_mea + 0.01


                print('mea_hou')
                print(mea_30)

                print('str_mea_num')
                print(str_mea_num)

                mea_30 = format(mea_30, f'.{(4 - str_mea_num)}f')



                print(mea_30)
                '''


        if num == '200' and mea == 'mA':

            if hz == 0:
                new_bzy.mea_dc('20mA')
                time.sleep(5)
                new_bzy.close_door()
                time.sleep(3)

        newItem = QTableWidgetItem('超量程')
        self.ui.run_table.setItem(11, lie, newItem)


        if '-' in num:

            len_value = len(num) - 1
            num = int(num)
            if len_value == 3:
                num = num + 1
                delta_bzy_mea = -1
            elif len_value == 2:
                num = num + 0.1
                delta_bzy_mea = -0.1
            elif len_value == 1:
                num = num + 0.01
                delta_bzy_mea = -0.01
            elif len_value == 4:
                num = num + 10
                delta_bzy_mea = -10
        else:
            len_value = len(num)
            num = int(num)
            if len_value == 3:
                num = num - 1
                delta_bzy_mea = 1
            elif len_value == 2:
                num = num - 0.1
                delta_bzy_mea = 0.1
            elif len_value == 1:
                num = num - 0.01
                delta_bzy_mea = 0.01
            elif len_value == 4:
                num = num - 10
                delta_bzy_mea = 10




        if hz == 0:
            mea_value = str(num) + mea
            print('测量单位')
            print(mea_value)
            bzy_mea = new_bzy.mea_dc(mea_value)
        elif hz != 0:
            mea_value = str(num) + mea
            print('hz单位')
            print(mea_value)
            bzy_mea = new_bzy.mea_ac(mea_value, hz)


        bzy_mea = float(bzy_mea)
        time.sleep(4)

        # 识别 lcd 数字
        photo_name = str(num) + mea + '_' + datetime.now().strftime("%Y%m%d%H%M%S")
        testGrabImage.paizhao(photo_name)
        time.sleep(1)
        self.run_image = "E:/photo/" + photo_name + ".jpg"

        mea_30 = ewm_and_lcd.lcd_shibie("E:/photo/" + photo_name + ".jpg")[0]


        # 处理 lcd 读数，加小数点，并显示于 ui 界面上
        mea_30 = int(mea_30)
        mea_30 = float(mea_30)

        xunhuan = 0

        while (True):

            print('bzy_mea')
            print(bzy_mea)
            value_test = mea_30 / bzy_mea

            if (value_test > 0.8 and value_test < 1.2) or (value_test < -0.8 and value_test > -1.2):

                break

            else:

                mea_30 = mea_30 * 0.1

                xunhuan = xunhuan + 1

                if xunhuan > 5:
                    print('sbsb')
                    mea_30 = mea_30 * 100000
                    break

        str_mea_num = len(str(int(bzy_mea)))

        if '-' in str(int(bzy_mea)):
            str_mea_num = str_mea_num - 1
            mea_30 = - mea_30

        print('mea_qian')
        print(mea_30)

        if delta_bzy_mea == -0.1:
            mea_30 = mea_30 - 0.1
            bzy_mea = bzy_mea - 0.1
        elif delta_bzy_mea == -1:
            mea_30 = mea_30 - 1
            bzy_mea = bzy_mea - 1
        elif delta_bzy_mea == -0.01:
            mea_30 = mea_30 - 0.01
            bzy_mea = bzy_mea - 0.01
        elif delta_bzy_mea == -10:
            mea_30 = mea_30 - 10
            bzy_mea = bzy_mea - 10
        elif delta_bzy_mea == 0.1:
            mea_30 = mea_30 + 0.1
            bzy_mea = bzy_mea + 0.1
        elif delta_bzy_mea == 1:
            mea_30 = mea_30 + 1
            bzy_mea = bzy_mea + 1
        elif delta_bzy_mea == 10:
            mea_30 = mea_30 + 10
            bzy_mea = bzy_mea + 10
        elif delta_bzy_mea == 0.01:
            mea_30 = mea_30 + 0.01
            bzy_mea = bzy_mea + 0.01

        print('mea_hou')
        print(mea_30)

        print('str_mea_num')
        print(str_mea_num)

        mea_30 = format(mea_30, f'.{(4 - str_mea_num)}f')








        '''
        if key == '4_2':
            mea_30 = str(mea_30 // 100) + '.' + str(mea_30 % 100)
        elif key == '3_2':
            mea_30 = str(mea_30 // 100) + '.' + str(mea_30 % 100).zfill(2)
        elif key == '4_1' or key == '3_1':
            mea_30 = str(mea_30 // 10) + '.' + str(mea_30 % 10)
            mea_30 = float(mea_30)
            if mea_30 < 1.0:
                mea_30 = mea_30 * 100
                mea_30 = str(mea_30)
                mea_30 = mea_30 + '.0'

        elif key == '4_3':
            mea_30 = str(f"{mea_30 / 1000:.3f}")
            mea_30 = float(mea_30)
            if mea_30 < 1:
                mea_30 = mea_30 * 1000
                mea_30 = str(mea_30)
        '''


        newItem = QTableWidgetItem(str(mea_30))
        self.ui.run_table.setItem(hang, lie, newItem)


        # 切换lcd图片
        self.alter_image()

        time.sleep(1)
        # 关闭标准源通道
        new_bzy.close_door()

        return mea_30, bzy_mea






    # 计算


    def caculator_num(self, vec_real, vec_mea, lie):

        print(vec_mea[0])
        point_num = int(self.count_decimal_places(vec_mea[0]))

        vec_mea[0] = float(vec_mea[0])
        vec_mea[1] = float(vec_mea[1])
        vec_mea[2] = float(vec_mea[2])


        self.ave_mea = round(((vec_mea[0] + vec_mea[1] + vec_mea[2]) / 3), point_num)
        self.max_mea = max([vec_mea[0], vec_mea[1], vec_mea[2]])
        self.min_mea = min([vec_mea[0], vec_mea[1], vec_mea[2]])
        self.vec_real_num = round(((vec_real[0] + vec_real[1] + vec_real[2]) / 3), 4)
        self.minus_mea = self.ave_mea - self.vec_real_num

        if self.ave_mea == 0 or self.vec_real_num == 0:

            self.rel_mea = 100.000
            self.rep_mea = 100.000

        else:


            self.rel_mea = float(format(((self.ave_mea - self.vec_real_num) / abs(self.vec_real_num)) * 100, f'.3f'))
            self.rep_mea = float(format(((self.max_mea - self.min_mea) / abs(self.ave_mea)) * 100, f'.3f'))


        if self.ui.run_table.item(11, lie) is not None:

            sql_result = 'overload'

        else:

            if self.rel_mea >= -0.5 and self.rel_mea <= 0.5:
                mea_result = '完美'
                sql_result = 'perfect'
                self.ui.run_table.setItem(11, lie, QTableWidgetItem(mea_result))
            elif self.rel_mea >= -1.0 and self.rel_mea <= 1.0:
                mea_result = '合格'
                sql_result = 'pass'
                self.ui.run_table.setItem(11, lie, QTableWidgetItem(mea_result))
            elif self.rel_mea >= -5.0 and self.rel_mea <= 5.0:
                mea_result = '不合格'
                sql_result = 'nopass'
                self.ui.run_table.setItem(11, lie, QTableWidgetItem(mea_result))
            else:
                mea_result = '数据异常'
                sql_result = 'unusual'
                self.ui.run_table.setItem(11, lie, QTableWidgetItem(mea_result))

        self.ave_mea = format(self.ave_mea, f'.{(point_num)}f')
        self.max_mea = format(self.max_mea, f'.{(point_num)}f')
        self.min_mea = format(self.min_mea, f'.{(point_num)}f')
        self.minus_mea = format(self.minus_mea, f'.{(point_num)}f')

        self.rel_mea = format(self.rel_mea, f'.3f') + '%'
        self.rep_mea = format(self.rep_mea, f'.3f') + '%'


        cursor = db.cursor()
        sql1 = "UPDATE wyb_value SET m_average = %s WHERE unit = %s AND measure = %s"
        cursor.execute(sql1, (str(self.ave_mea), vec_mea[3], str(vec_mea[4])))
        db.commit()
        sql1 = "UPDATE wyb_value SET m_max = %s WHERE unit = %s AND measure = %s"
        cursor.execute(sql1, (str(self.max_mea), vec_mea[3], str(vec_mea[4])))
        db.commit()
        sql1 = "UPDATE wyb_value SET m_max = %s WHERE unit = %s AND measure = %s"
        cursor.execute(sql1, (str(self.max_mea), vec_mea[3], str(vec_mea[4])))
        db.commit()
        sql1 = "UPDATE wyb_value SET m_min = %s WHERE unit = %s AND measure = %s"
        cursor.execute(sql1, (str(self.min_mea), vec_mea[3], str(vec_mea[4])))
        db.commit()
        sql1 = "UPDATE wyb_value SET m_sz = %s WHERE unit = %s AND measure = %s"
        cursor.execute(sql1, (str(self.minus_mea), vec_mea[3], str(vec_mea[4])))
        db.commit()
        sql1 = "UPDATE wyb_value SET m_xd = %s WHERE unit = %s AND measure = %s"
        cursor.execute(sql1, (str(self.rel_mea), vec_mea[3], str(vec_mea[4])))
        db.commit()
        sql1 = "UPDATE wyb_value SET m_repeat = %s WHERE unit = %s AND measure = %s"
        cursor.execute(sql1, (str(self.rep_mea), vec_mea[3], str(vec_mea[4])))
        db.commit()


        sql1 = "UPDATE wyb_value SET pass = %s WHERE unit = %s AND measure = %s"
        cursor.execute(sql1, (str(sql_result), vec_mea[3], str(vec_mea[4])))
        db.commit()





        self.ui.run_table.setItem(5, lie, QTableWidgetItem(str(self.ave_mea)))
        self.ui.run_table.setItem(6, lie, QTableWidgetItem(str(self.max_mea)))
        self.ui.run_table.setItem(7, lie, QTableWidgetItem(str(self.min_mea)))
        self.ui.run_table.setItem(8, lie, QTableWidgetItem(str(self.minus_mea)))
        self.ui.run_table.setItem(9, lie, QTableWidgetItem(self.rel_mea))
        self.ui.run_table.setItem(10, lie, QTableWidgetItem(self.rep_mea))










    # 其他


    def alter_image(self):

        self.pixmap = QPixmap(self.run_image)
        self.new_pixmap = self.pixmap.scaled(self.ui.run_graph.width(), self.ui.run_graph.height())
        self.ui.run_graph.setPixmap(self.new_pixmap)


    def run_clear_wyb(self):

        for i in range(39):

            for j in range(11):

                self.run_table.setItem(j, i + 1, None)

        self.ui.run_wyb_label_unit.setText(' ')
        self.ui.run_wyb_label_value.setText(' ')
        self.ui.run_wyb_label_result.setText(' ')


    def run_exit(self):

        self.zhujiemian = wyb()
        self.zhujiemian.ui.show()
        self.ui.close()


    def run_stop(self):

        pid = os.getpid()  # 获取当前进程的PID
        os.kill(pid, signal.SIGTERM)  # 发
        robot_act_new.sys.exit("1")



    def count_decimal_places(self,num):
        # 将数字转换为字符串

        str_num = str(num)
        # 找到小数点的位置
        decimal_point_index = str_num.find('.')
        # 如果没有小数点，返回0
        if decimal_point_index == -1:
            return 0
        # 计算小数点后的字符数量
        decimal_places = len(str_num) - decimal_point_index - 1
        return decimal_places




class run_manual():


    def __init__(self):

        self.ui = QUiLoader().load(addr_ui + "/run_manual.ui")

        self.ui.run_start.setVisible(True)

        self.ui.run_exit.clicked.connect(self.run_exit)  # 下一步

        self.ui.run_stop.clicked.connect(self.run_stop)  # 下一步

        self.ui.run_start.clicked.connect(self.run_start_system)

        self.ui.run_go_on.clicked.connect(self.run_go_on)

        self.ui.run_start.setStyleSheet("background-color: yellow;")
        self.ui.run_stop.setStyleSheet("background-color: yellow;color: red;")
        self.ui.run_wyb_num.setStyleSheet("color: blue;")
        self.ui.run_wyb_nio.setStyleSheet("color: blue;")
        self.ui.run_wyb_type.setStyleSheet("color: blue;")

        self.key = 0

        self.run_table = self.ui.run_table

        self.run_clear_wyb()

        self.run_show_wyb_info(1)

        self.cha_kong = False

        self.run_image = 'E:\photo_test\lcd_test.jpg'

        self.alter_image()

        self.ui.run_wyb_message.setText("请点击开始测量！")

        self.ui.run_wyb_message.setStyleSheet(
            '''color: red; justify-content: center; align-items: center; text-align: center;''')

        cursor = db.cursor()

        # 查 最大装夹数
        cursor.execute('SELECT max(equip_num) FROM wyb_test_info')
        db.commit()
        max_equip_num = int(cursor.fetchone()[0])

        for select_num in range(1, max_equip_num + 1):

            self.ui.run_wyb_select.addItem(str(select_num))




    # 前期验证 + 显示信息

    def run_show_wyb_info(self, wyb_work_num):

        global current_wyb

        current_wyb = wyb_work_num

        # 一、更新屏幕万用表信息
        wyb_info_num = '第 ' + str(wyb_work_num) + ' 台'
        self.ui.run_wyb_num.setText(wyb_info_num)

        wyb_work_num = str(wyb_work_num)

        cursor = db.cursor()
        sql = "SELECT num,nio FROM wyb_test_info WHERE equip_num = %s"
        cursor.execute(sql, wyb_work_num)
        db.commit()
        result = cursor.fetchone()

        wyb_num = str(result[0])
        wyb_nio = str(result[1])

        sql = "SELECT type FROM wyb_name_info WHERE num = %s"
        cursor.execute(sql, wyb_num)
        db.commit()
        result = cursor.fetchone()
        wyb_type = str(result[0])

        self.ui.run_wyb_type.setText(wyb_num)
        self.ui.run_wyb_nio.setText(wyb_nio)

        wyb_pos_add = "E:/photo_test/" + wyb_type + "_pos.jpg"

        # print(wyb_pos_add)

        self.alter_kong_image(wyb_pos_add)

        testGrabImage.paizhao('1')

        # 二、万用表数据库

        sql = "SELECT unit FROM wyb_value WHERE num = %s"
        cursor.execute(sql, wyb_num)
        db.commit()
        wyb_unit = cursor.fetchall()

        sql = "SELECT measure FROM wyb_value WHERE num = %s"
        cursor.execute(sql, wyb_num)
        db.commit()
        wyb_measure = cursor.fetchall()

        length1 = len(wyb_measure)

        for i in range(0, length1):
            v = str(wyb_unit[i][0])

            if "R" in v or "KR" in v or "MR" in v:
                # 直接使用原始值，因为这些值不需要转换
                self.ui.run_table.setItem(0, i + 1, QTableWidgetItem(v))
                self.ui.run_table.setItem(1, i + 1, QTableWidgetItem(str(wyb_measure[i][0])))
            else:
                # 检查是否包含 "AC" 或 "DC"
                if "AC" in v:
                    # 处理包含 "AC" 的情况
                    parts = v.split('AC')
                    sign = "AC"  # 正负号为 "AC"
                elif "DC" in v:
                    # 处理包含 "DC" 的情况
                    parts = v.split('DC')
                    sign = "DC"

                old_value = parts[0].strip() + parts[1].strip()
                value = ''.join([char.lower() if char in ['M', 'U'] else char for char in old_value])

                sign1 = ''

                if '+' in value:
                    sign1 = '+'
                    value = value.replace('+', '')  # 移除符号
                elif '-' in value:
                    sign1 = '-'
                    value = value.replace('-', '')  # 移除符号

                sign += sign1

                # 构建新的字符串
                new_value = f"{value}({sign})"
                self.ui.run_table.setItem(0, i + 1, QTableWidgetItem(new_value))
                self.ui.run_table.setItem(1, i + 1, QTableWidgetItem(str(wyb_measure[i][0])))

    def run_modify_work_start_wyb_num(self,sel):

        self.P_begin_height = copy.copy(robot_act_new.P_begin)

        self.P_begin_height[2] = -80.000

        robot_act_new.movej_PTP_getJ(3, 70.0, self.P_begin_height)

        robot_act_new.movej_PTP_getJ(3, 70.0, robot_act_new.P_begin)

        time.sleep(3)

        # 1、拍当前装夹位，计算得到转盘转动格数，使 工作位 转至 1 号 位

        photo_name = "robot_equip_pic"
        testGrabImage.paizhao(photo_name)
        time.sleep(5)
        testGrabImage.paizhao(photo_name)
        self.run_image = 'E:/photo/robot_equip_pic.jpg'
        self.alter_image()
        time.sleep(2)
        dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg")
        self.run_image = 'E:/photo/ewm_status.jpg'
        self.alter_image()

        if dedaozhi != 'shibai':

            jiaodu = 110
            self.ui.run_wyb_message.setText("扫码成功，角度：110！")
            self.run_image = 'E:/photo/ewm_status.jpg'
            self.alter_image()
            pass

        elif dedaozhi == 'shibai':

            self.ui.run_wyb_message.setText("扫码失败，调整角度至120！")
            time.sleep(3)
            testGrabImage.paizhao(photo_name)
            time.sleep(2)
            jiaodu = 120
            dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg", jiaodu)
            self.run_image = 'E:/photo/ewm_status.jpg'
            self.alter_image()
            if dedaozhi == 'shibai':
                self.ui.run_wyb_message.setText("扫码失败，调整角度至130！")
                time.sleep(3)
                testGrabImage.paizhao(photo_name)
                time.sleep(2)
                jiaodu = 130
                dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg", jiaodu)
                self.run_image = 'E:/photo/ewm_status.jpg'
                self.alter_image()
                if dedaozhi == 'shibai':
                    self.ui.run_wyb_message.setText("扫码失败，调整角度至135！")
                    time.sleep(3)
                    testGrabImage.paizhao(photo_name)
                    time.sleep(2)
                    jiaodu = 135
                    dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg", jiaodu)
                    self.run_image = 'E:/photo/ewm_status.jpg'
                    self.alter_image()
                    if dedaozhi == 'shibai':
                        self.ui.run_wyb_message.setText("扫码失败，调整角度至100！")
                        time.sleep(3)
                        testGrabImage.paizhao(photo_name)
                        time.sleep(2)
                        jiaodu = 100
                        dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg", jiaodu)
                        self.run_image = 'E:/photo/ewm_status.jpg'
                        self.alter_image()
                        if dedaozhi == 'shibai':
                            self.ui.run_wyb_message.setText("扫码失败，调整角度至90！")
                            time.sleep(3)
                            testGrabImage.paizhao(photo_name)
                            time.sleep(2)
                            jiaodu = 90
                            dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg", jiaodu)
                            self.run_image = 'E:/photo/ewm_status.jpg'
                            self.alter_image()

        if dedaozhi == 'shibai':

            self.ui.run_wyb_message.setText("扫码失败!")
            self.run_image = 'E:/photo/ewm_status.jpg'
            self.alter_image()
            QMessageBox.information(QMessageBox(), '验证', '扫码失败，请手动重启程序，至当前页面重新开始测量！')

            equip_num_again = -1

        else:

            text_1 = "角度：" + str(jiaodu) + " 转大转盘中："

            self.ui.run_wyb_message.setText(text_1)

            equip_num = int(dedaozhi)


            if int(sel) == 1:

                if equip_num in [1, 2, 3]:

                    robot_act_new.send_zp_start(4 - equip_num)

                elif equip_num in [5, 6, 7, 8]:

                    robot_act_new.send_zp_start(12 - equip_num)

            elif int(sel) == 2:

                if equip_num in [1, 2, 3, 4, 5]:

                    robot_act_new.send_zp_start(5 - equip_num)

                elif equip_num in [6, 7, 8]:

                    robot_act_new.send_zp_start(13 - equip_num)

            elif int(sel) == 3:


                if equip_num in [1, 2, 3, 4, 5, 6]:

                    robot_act_new.send_zp_start(6 - equip_num)

                elif equip_num in [7, 8]:

                    robot_act_new.send_zp_start(14 - equip_num)

            elif int(sel) == 4:

                if equip_num in [1, 2, 3, 4, 5, 6, 7]:

                    robot_act_new.send_zp_start(7 - equip_num)

                elif equip_num in [8]:

                    robot_act_new.send_zp_start(15 - equip_num)

            elif int(sel) == 5:

                robot_act_new.send_zp_start(8 - equip_num)

            elif int(sel) == 6:

                if equip_num in [1]:

                    robot_act_new.send_zp_start(1 - equip_num)

                elif equip_num in [2, 3, 4, 5, 6, 7, 8]:

                    robot_act_new.send_zp_start(9 - equip_num)

            elif int(sel) == 7:

                if equip_num in [1, 2]:

                    robot_act_new.send_zp_start(2 - equip_num)

                elif equip_num in [3, 4, 5, 6, 7, 8]:

                    robot_act_new.send_zp_start(10 - equip_num)

            elif int(sel) == 8:

                if equip_num in [1, 2, 3]:

                    robot_act_new.send_zp_start(3 - equip_num)

                elif equip_num in [4, 5, 6, 7, 8]:

                    robot_act_new.send_zp_start(11 - equip_num)



            # 2、验证 工作位 为 1 号 位（装夹位 为 4 号 位）

            self.ui.run_wyb_message.setText("再次验证二维码！")

            time.sleep(5)

            photo_name_again = "robot_equip_pic_again"
            testGrabImage.paizhao(photo_name_again)
            time.sleep(2)
            dedaozhi_again = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name_again + ".jpg")
            self.run_image = 'E:/photo/ewm_status.jpg'
            self.alter_image()
            if dedaozhi_again != 'shibai':
                self.ui.run_wyb_message.setText("验证扫码成功！验证角度：110")
                self.run_image = 'E:/photo/ewm_status.jpg'
                self.alter_image()
                pass
            elif dedaozhi_again == 'shibai':
                self.ui.run_wyb_message.setText("验证扫码失败，调整角度至120！")
                time.sleep(3)
                testGrabImage.paizhao(photo_name_again)
                time.sleep(2)
                jiaodu = 120
                dedaozhi_again = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name_again + ".jpg", jiaodu)
                self.run_image = 'E:/photo/ewm_status.jpg'
                self.alter_image()
                if dedaozhi_again == 'shibai':
                    self.ui.run_wyb_message.setText("验证扫码失败，调整角度至140！")
                    time.sleep(3)
                    testGrabImage.paizhao(photo_name_again)
                    time.sleep(2)
                    jiaodu = 140
                    dedaozhi_again = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name_again + ".jpg", jiaodu)
                    self.run_image = 'E:/photo/ewm_status.jpg'
                    self.alter_image()
                    if dedaozhi_again == 'shibai':
                        self.ui.run_wyb_message.setText("验证扫码失败，调整角度至100！")
                        time.sleep(3)
                        testGrabImage.paizhao(photo_name_again)
                        time.sleep(2)
                        jiaodu = 100
                        dedaozhi_again = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name_again + ".jpg", jiaodu)
                        self.run_image = 'E:/photo/ewm_status.jpg'
                        self.alter_image()
                        if dedaozhi_again == 'shibai':
                            self.ui.run_wyb_message.setText("验证扫码失败，调整角度至130！")
                            time.sleep(3)
                            testGrabImage.paizhao(photo_name_again)
                            time.sleep(2)
                            jiaodu = 130
                            dedaozhi_again = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name_again + ".jpg", jiaodu)
                            self.run_image = 'E:/photo/ewm_status.jpg'
                            self.alter_image()

            if dedaozhi_again == 'shibai':
                self.ui.run_wyb_message.setText("验证扫码失败！")
                self.run_image = 'E:/photo/ewm_status.jpg'
                self.alter_image()

                QMessageBox.information(QMessageBox(), '验证', '扫码失败，请手动重启程序，至当前页面重新开始测量！')

                equip_num_again = -1

            else:
                text_1 = "验证扫码成功！验证角度：" + str(jiaodu)
                self.ui.run_wyb_message.setText(text_1)
                equip_num_again = int(dedaozhi_again)

                print('第二次验证值：' + str(equip_num_again))


        return equip_num_again

    def run_find_kong_zp_pos(self, type, wyb_num):

        global tool_x, tool_y

        cursor = db.cursor()

        cursor.execute("SELECT d_all_wyb FROM jiajv_info WHERE id = 1")
        db.commit()
        result = cursor.fetchone()
        x1, y1, z1, rx1, ry1, rz1 = map(float, result[0].split('&'))
        W_photo = [x1, y1, z1, rx1, ry1, rz1]

        robot_act_new.movej_PTP_getJ(3, 60.0, W_photo)

        self.ui.run_wyb_message.setText("正识别万用表拨盘位置！")

        time.sleep(2)

        testGrabImage.paizhao(wyb_num, 0)

        time.sleep(2)

        testGrabImage.paizhao(wyb_num, 0)

        convert_new.pic_equip_path = "E:/photo_test/" + wyb_num + ".jpg"
        convert_new.pic_blue_path = "E:/photo_test/" + type + "_ori.jpg"

        time.sleep(1)

        # 0.241935


        wyb_all_loc,kong_loc,res,zp_loc = convert_new.get_now_wyb_loc(4, 0.243)

        if kong_loc is None:

            print('第一次标定万用表失败，第二次标定')

            testGrabImage.paizhao(wyb_num, 0)

            time.sleep(2.5)

            testGrabImage.paizhao(wyb_num, 0)

            convert_new.pic_equip_path = "E:/photo_test/" + wyb_num + ".jpg"
            convert_new.pic_blue_path = "E:/photo_test/" + type + "_ori.jpg"

            time.sleep(4)

            wyb_all_loc, kong_loc, res, zp_loc = convert_new.get_now_wyb_loc(4, 0.243)

            if kong_loc is None:

                print('第二次标定万用表失败，第三次标定')

                testGrabImage.paizhao(wyb_num, 0)

                time.sleep(2.5)

                testGrabImage.paizhao(wyb_num, 0)

                convert_new.pic_equip_path = "E:/photo_test/" + wyb_num + ".jpg"
                convert_new.pic_blue_path = "E:/photo_test/" + type + "_ori.jpg"

                time.sleep(1)

                wyb_all_loc, kong_loc, res, zp_loc = convert_new.get_now_wyb_loc(4, 0.243)


                if kong_loc is None:

                    self.ui.run_wyb_message.setText("识别失败，请重启程序选择该台！")

                    return -1


        cv2.imwrite('E:/photo_test/res.jpg', res)

        k1_loc_x = x1 - kong_loc[0][1] + tool_x
        k1_loc_y = kong_loc[0][0] + y1 + tool_y
        k2_loc_x = x1 - kong_loc[1][1] + tool_x
        k2_loc_y = kong_loc[1][0] + y1 + tool_y
        k3_loc_x = x1 - kong_loc[2][1] + tool_x
        k3_loc_y = kong_loc[2][0] + y1 + tool_y
        k4_loc_x = x1 - kong_loc[3][1] + tool_x
        k4_loc_y = kong_loc[3][0] + y1 + tool_y

        zp_loc_x = x1 - zp_loc[1] + tool_x + buchang_x
        zp_loc_y = zp_loc[0] + y1 + tool_y + buchang_y

        #print('y1')
        #print(y1)
        #print('kong_loc')
        #print(kong_loc[0][0])

        k1_loc_x = round(k1_loc_x, 3)
        k1_loc_y = round(k1_loc_y, 3)
        k2_loc_x = round(k2_loc_x, 3)
        k2_loc_y = round(k2_loc_y, 3)
        k3_loc_x = round(k3_loc_x, 3)
        k3_loc_y = round(k3_loc_y, 3)
        k4_loc_x = round(k4_loc_x, 3)
        k4_loc_y = round(k4_loc_y, 3)

        zp_loc_x = round(zp_loc_x, 3)
        zp_loc_y = round(zp_loc_y, 3)

        k1_pos = f"{k1_loc_x}&{k1_loc_y}"
        k2_pos = f"{k2_loc_x}&{k2_loc_y}"
        k3_pos = f"{k3_loc_x}&{k3_loc_y}"
        k4_pos = f"{k4_loc_x}&{k4_loc_y}"
        zp_pos = f"{zp_loc_x}&{zp_loc_y}"

        sql = "UPDATE wyb_info SET k2_pos = %s WHERE type = %s"
        cursor.execute(sql, (k1_pos, type))
        db.commit()
        sql = "UPDATE wyb_info SET k4_pos = %s WHERE type = %s"
        cursor.execute(sql, (k2_pos, type))
        db.commit()
        sql = "UPDATE wyb_info SET k1_pos = %s WHERE type = %s"
        cursor.execute(sql, (k3_pos, type))
        db.commit()
        sql = "UPDATE wyb_info SET k3_pos = %s WHERE type = %s"
        cursor.execute(sql, (k4_pos, type))
        db.commit()

        sql = "UPDATE wyb_info SET zp_pos = %s WHERE type = %s"
        cursor.execute(sql, (zp_pos, type))
        db.commit()

        print('全景图拍摄验证成功！')
        print(zp_pos)



    # 插针还针
    def run_go_on(self):

        self.cha_kong = True



    # 目前不用（验证下一台万用表 测量位的二维码）
    def run_modify_work_next_wyb_num(self, modify_erm_num):

        self.P_modify_height = copy.copy(robot_act_new.P_modify)

        self.P_modify_height[2] = -80.000

        robot_act_new.movej_PTP_getJ(3, 70.0, self.P_modify_height)

        robot_act_new.movej_PTP_getJ(3, 70.0, robot_act_new.P_modify)

        time.sleep(3)

        # 扫码出图
        photo_name = "modify_work_ewm"
        testGrabImage.paizhao(photo_name)
        self.run_image = 'E:/photo/modify_work_ewm.jpg'
        self.alter_image()
        time.sleep(2)
        dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg")
        if dedaozhi != 'shibai':
            equip_num = int(dedaozhi)
            self.run_image = 'E:/photo/ewm_status.jpg'
            self.alter_image()
        elif dedaozhi == 'shibai':
            time.sleep(3)
            testGrabImage.paizhao(photo_name)
            time.sleep(2)
            dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg", 120)
            self.run_image = 'E:/photo/ewm_status.jpg'
            self.alter_image()
            if dedaozhi == 'shibai':
                time.sleep(3)
                testGrabImage.paizhao(photo_name)
                time.sleep(2)
                dedaozhi = ewm_and_lcd.erweima_shibei("E:/photo/" + photo_name + ".jpg", 140)
                self.run_image = 'E:/photo/ewm_status.jpg'
                self.alter_image()

        equip_num = int(dedaozhi)

        print('得到值：')
        print(equip_num)
        if equip_num != modify_erm_num:

            return False

        elif equip_num == modify_erm_num:

            return True



    # 开始运行



    # 开始测量 按钮，开线程（ewm_func）
    def run_start_system(self):

        reply = QMessageBox.question(QMessageBox(), '运行', '请确认将双门关闭！开始后，则删除全部照片！并检查左上角即将运行的序号！（可取消）',
                                     QMessageBox.Yes | QMessageBox.No, QMessageBox.No)

        if reply == QMessageBox.Yes:


            self.equip_num_again = 0

            time.sleep(2)

            wyb_modify_ewm1 = threading.Thread(target=self.wyb_modify_ewm_func)

            wyb_modify_ewm1.start()

        else:

            return


    # 扫工作位二维码，随后开线程（start_wyb），连续测量八台
    def wyb_modify_ewm_func(self):

        print('清空照片')

        base_folder_path = "E:/photo"
        # 遍历p1到p8的文件夹
        for i in range(1, 9):  # 从1到8
            folder_path = os.path.join(base_folder_path, f"p{i}")  # 构建每个子文件夹的路径
            # 检查文件夹是否存在
            if os.path.exists(folder_path) and os.path.isdir(folder_path):
                # 遍历文件夹中的文件
                for filename in os.listdir(folder_path):
                    # 检查文件扩展名是否在图片扩展名列表中
                    if os.path.splitext(filename)[1].lower() in ['.jpg']:
                        # 构建完整的文件路径
                        file_path = os.path.join(folder_path, filename)
                        # 删除文件
                        os.remove(file_path)

        print('线程启动，开始测量！')

        sel_num = str(self.ui.run_wyb_select.currentText())

        # 移动机械臂至安全高度
        robot_act_new.back_to_safe_plane()

        self.ui.run_start.setVisible(False)

        self.ui.run_wyb_message.setText("移动机械臂至扫码位！")

        # 复位 和 验证

        self.equip_num_again = self.run_modify_work_start_wyb_num(sel_num)


        #self.equip_num_again = 4


        if self.equip_num_again == 411:

            QMessageBox.information(QMessageBox(), '运行', '目前工位不为 1号工位，扫码失败，请重启程序！')

        elif self.equip_num_again != 411:

            self.run_clear_wyb()

            self.run_show_wyb_info(sel_num)

            wyb_system_start = threading.Thread(target=self.run_start_wyb)

            print('开始运行！')

            wyb_system_start.start()

    # 连测八台
    def run_start_wyb(self):

        global k_num, be, en, photo_num

        cursor = db.cursor()

        # 查 最大装夹数
        cursor.execute('SELECT max(equip_num) FROM wyb_test_info')
        db.commit()
        max_equip_num = int(cursor.fetchone()[0])

        start_wyb_num = int(self.ui.run_wyb_select.currentText())

        for equip_num in range(start_wyb_num, max_equip_num + 1):

            # 一、查询 万用表 信息

            photo_num = equip_num

            # 查 内码 num
            cursor.execute('SELECT num FROM wyb_test_info WHERE equip_num = %s', (equip_num,))
            db.commit()
            wyb_num = cursor.fetchone()[0]

            # 查 型号 type
            cursor.execute('SELECT type FROM wyb_name_info WHERE num = %s', (wyb_num,))
            db.commit()
            wyb_type = cursor.fetchone()[0]

            # 二、拍照——————识别——————4孔+拨盘中心

            time.sleep(3)

            self.ui.run_wyb_message.setText("正移动机械臂至工作位！")

            _return = self.run_find_kong_zp_pos(wyb_type, wyb_num)

            if _return == -1:
                return

            # 三、 关闭 标准源 通道
            self.ui.run_wyb_message.setText("关闭标准源通道中！")
            #new_bzy.close_door()

            # 四、 手动插针

            time.sleep(1)

            self.ui.run_wyb_message.setText("请插针！")
            self.ui.run_wyb_message.setStyleSheet(
                '''color: red; justify-content: center; align-items: center; text-align: center;''')

            robot_act_new.send_voice()

            self.cha_kong = False

            while (True):
                if self.cha_kong is True:
                    break
                pass

            self.cha_kong = False

            time.sleep(1)


            # 五、 测量 第一台 万用表 FLUKE

            self.ui.run_wyb_message.setText("测量档位中！")
            self.run_mea_vc(36, wyb_num, wyb_type)

            # 六、  还针
            self.cha_kong = False
            time.sleep(2)


            self.ui.run_wyb_message.setText("请还针！")
            robot_act_new.send_voice()

            while (True):
                if self.cha_kong is True:
                    break
                pass

            self.cha_kong = False

            time.sleep(2)

            print(self.cha_kong)

            # 七、 转动转盘，清空表格数据，显示下一台数据，初始化参数 k_num,be,en

            if equip_num != max_equip_num:

                # 验证转盘
                self.ui.run_wyb_message.setText("转动转盘中！")
                robot_act_new.send_zp_start(1)

                self.ui.run_wyb_message.setText("验证当前装夹位中！")
                time.sleep(2)

                modify_key = True

                # 验证 下一台 程序
                '''
                if equip_num >= 1 and equip_num <= 6:

                    modify_key = self.run_modify_work_next_wyb_num(equip_num + 2)

                    time.sleep(2)


                else:

                    modify_key = self.run_modify_work_next_wyb_num(equip_num - 6)
                '''

                if modify_key is False:

                    QMessageBox.information(QMessageBox(), '测量', '当前转盘出现问题！')

                    return

                elif modify_key is True:

                    # QMessageBox.information(QMessageBox(), '测量', '2号万用表开始测量！')

                    self.run_clear_wyb()

                    self.run_show_wyb_info(equip_num + 1)

                    self.ui.run_wyb_message.setText("下一台开始！")

                    # 将 计数参数 初始化

                    self.cha_kong = False
                    print(self.cha_kong)

                    k_num = 2
                    be = 0
                    en = 1

                    time.sleep(2)

        self.ui.run_wyb_message.setText("所有万用表测量完毕！！")

    # 万用表不同型号
    def run_mea_vc(self, all_mea_num, wyb_num, wyb_type):

        global k_num, be, en

        cursor = db.cursor()

        #  先测电流

        unit1 = 'OFF'


        for i in range(1, all_mea_num + 1):

            print('新测点开始：')

            print('上一个档位：' + unit1)

            sql = "SELECT dang_angle FROM wyb_unit_info WHERE dang_type = %s and type = %s"
            cursor.execute(sql, (unit1, wyb_type))
            db.commit()
            result = cursor.fetchone()

            begin_angle = float(result[0])

            sql = "SELECT unit FROM wyb_value WHERE id_num = %s and num = %s"
            cursor.execute(sql, (i, wyb_num))
            db.commit()
            result = cursor.fetchone()
            unit1 = str(result[0])

            print('新档位：' + unit1)

            unit1 = unit1.replace('+', '').replace('-', '')


            sql = "SELECT dang_angle FROM wyb_unit_info WHERE dang_type = %s and type = %s"
            cursor.execute(sql, (unit1, wyb_type))
            db.commit()
            result = cursor.fetchone()

            end_angle = float(result[0])

            # print(begin_angle + '//' + end_angle)

            if (begin_angle == end_angle):

                pass

                print('不转')

            else:

                robot_act_new.execute_robot_rotate(begin_angle, end_angle, wyb_type)

                print('该档位需要旋转拨盘')

            time.sleep(2)



            # 去掉了测量，只验证转盘


            
            #  四、查询 万用表 测点等信息，通过函数 转换为 标准源 可读 测点

            sql = "SELECT unit,measure,hz FROM wyb_value WHERE id_num = %s and num = %s"

            cursor.execute(sql, (i, wyb_num))
            db.commit()
            result = cursor.fetchone()
            show_unit = str(result[0])
            mea_value = str(result[1])
            hz_value = int(result[2])

            mea_unit = self.run_check_num(show_unit)

            #print(mea_value)

            #  五、 测量 该台 万用表

            if hz_value != 0:

                self.run_mea_same(show_unit, mea_unit, mea_value, vector_v1, bzy_v1, i, hz_value, wyb_num)

            else:

                self.run_mea_same(show_unit, mea_unit, mea_value, vector_v1, bzy_v1, i, 0, wyb_num)

            # 六、保存测量值，测量单位等信息，计算公式，随后导入数据库

            vector_v1[3] = show_unit
            vector_v1[4] = mea_value

            self.caculator_num(bzy_v1, vector_v1, i, wyb_num)
            



        '''
        for i in range(1, 26 + 1):

            print('上一个档位：' + unit1)

            sql = "SELECT dang_angle FROM wyb_unit_info WHERE dang_type = %s and type = %s"
            cursor.execute(sql, (unit1, wyb_type))
            db.commit()
            result = cursor.fetchone()

            begin_angle = float(result[0])

            sql = "SELECT unit FROM wyb_value WHERE id_num = %s and num = %s"
            cursor.execute(sql, (i, wyb_num))
            db.commit()
            result = cursor.fetchone()
            unit1 = str(result[0])

            print('处理前：' + 'unit1')

            unit1 = unit1.replace('+', '').replace('-', '')

            print('处理后：' + 'unit1')

            sql = "SELECT dang_angle FROM wyb_unit_info WHERE dang_type = %s and type = %s"
            cursor.execute(sql, (unit1, wyb_type))
            db.commit()
            result = cursor.fetchone()

            end_angle = float(result[0])

            #print(begin_angle + '//' + end_angle)

            if (begin_angle == end_angle):

                pass

                print('不转')

            else:

                robot_act_new.execute_robot_rotate(begin_angle, end_angle, wyb_type)

                print('转')

            time.sleep(2)

            
            #  四、查询 万用表 测点等信息，通过函数 转换为 标准源 可读 测点

            sql = "SELECT unit,measure,hz FROM wyb_value WHERE id_num = %s and num = %s"

            cursor.execute(sql, (i, wyb_num))
            db.commit()
            result = cursor.fetchone()
            show_unit = str(result[0])
            mea_value = str(result[1])
            hz_value = int(result[2])

            mea_unit = self.run_check_num(show_unit)

            print(mea_value)

            #  五、 测量 该台 万用表

            if hz_value != 0:

                self.run_mea_same(show_unit, mea_unit, mea_value, vector_v1, bzy_v1, i, hz_value, wyb_num)

            else:

                self.run_mea_same(show_unit, mea_unit, mea_value, vector_v1, bzy_v1, i, 0, wyb_num)

            # 六、保存测量值，测量单位等信息，计算公式，随后导入数据库

            vector_v1[3] = show_unit
            vector_v1[4] = mea_value

            self.caculator_num(bzy_v1, vector_v1, i, wyb_num)

            
        '''



        # 正式如下

        '''
                for i in range(15, all_mea_num + 1):

            print('上一个档位：' + unit1)

            sql = "SELECT dang_angle FROM wyb_unit_info WHERE dang_type = %s and type = %s"
            cursor.execute(sql, (unit1, wyb_type))
            db.commit()
            result = cursor.fetchone()

            begin_angle = float(result[0])

            sql = "SELECT unit FROM wyb_value WHERE id_num = %s and num = %s"
            cursor.execute(sql, (i, wyb_num))
            db.commit()
            result = cursor.fetchone()
            unit1 = str(result[0])

            print('处理前：' + 'unit1')

            unit1 = unit1.replace('+', '').replace('-', '')

            print('处理后：' + 'unit1')

            sql = "SELECT dang_angle FROM wyb_unit_info WHERE dang_type = %s and type = %s"
            cursor.execute(sql, (unit1, wyb_type))
            db.commit()
            result = cursor.fetchone()

            end_angle = float(result[0])

            # print(begin_angle + '//' + end_angle)

            if (begin_angle == end_angle):

                pass

                print('不转')

            else:

                robot_act_new.execute_robot_rotate(begin_angle, end_angle, wyb_type)

                print('转')

            #  四、查询 万用表 测点等信息，通过函数 转换为 标准源 可读 测点

            sql = "SELECT unit,measure,hz FROM wyb_value WHERE id_num = %s and num = %s"

            cursor.execute(sql, (i, wyb_num))
            db.commit()
            result = cursor.fetchone()
            show_unit = str(result[0])
            mea_value = str(result[1])
            hz_value = int(result[2])

            mea_unit = self.run_check_num(show_unit)

            print(mea_value)

            #  五、 测量 该台 万用表

            if hz_value != 0:

                self.run_mea_same(show_unit, mea_unit, mea_value, vector_v1, bzy_v1, i, hz_value, wyb_num)

            else:

                self.run_mea_same(show_unit, mea_unit, mea_value, vector_v1, bzy_v1, i, 0, wyb_num)

            # 六、保存测量值，测量单位等信息，计算公式，随后导入数据库

            vector_v1[3] = show_unit
            vector_v1[4] = mea_value

            self.caculator_num(bzy_v1, vector_v1, i, wyb_num)

        for i in range(1, 14 + 1):

            print('上一个档位：' + unit1)

            sql = "SELECT dang_angle FROM wyb_unit_info WHERE dang_type = %s and type = %s"
            cursor.execute(sql, (unit1, wyb_type))
            db.commit()
            result = cursor.fetchone()

            begin_angle = float(result[0])

            sql = "SELECT unit FROM wyb_value WHERE id_num = %s and num = %s"
            cursor.execute(sql, (i, wyb_num))
            db.commit()
            result = cursor.fetchone()
            unit1 = str(result[0])

            print('处理前：' + 'unit1')

            unit1 = unit1.replace('+', '').replace('-', '')

            print('处理后：' + 'unit1')

            sql = "SELECT dang_angle FROM wyb_unit_info WHERE dang_type = %s and type = %s"
            cursor.execute(sql, (unit1, wyb_type))
            db.commit()
            result = cursor.fetchone()

            end_angle = float(result[0])

            #print(begin_angle + '//' + end_angle)

            if (begin_angle == end_angle):

                pass

                print('不转')

            else:

                robot_act_new.execute_robot_rotate(begin_angle, end_angle, wyb_type)

                print('转')

            #  四、查询 万用表 测点等信息，通过函数 转换为 标准源 可读 测点

            sql = "SELECT unit,measure,hz FROM wyb_value WHERE id_num = %s and num = %s"

            cursor.execute(sql, (i, wyb_num))
            db.commit()
            result = cursor.fetchone()
            show_unit = str(result[0])
            mea_value = str(result[1])
            hz_value = int(result[2])

            mea_unit = self.run_check_num(show_unit)

            print(mea_value)

            #  五、 测量 该台 万用表

            if hz_value != 0:

                self.run_mea_same(show_unit, mea_unit, mea_value, vector_v1, bzy_v1, i, hz_value, wyb_num)

            else:

                self.run_mea_same(show_unit, mea_unit, mea_value, vector_v1, bzy_v1, i, 0, wyb_num)

            # 六、保存测量值，测量单位等信息，计算公式，随后导入数据库

            vector_v1[3] = show_unit
            vector_v1[4] = mea_value

            self.caculator_num(bzy_v1, vector_v1, i, wyb_num)
        '''


        #  正常测量

        '''

        unit1 = 'OFF'

        for i in range(1, all_mea_num + 1):


            print('上一个档位：' + unit1)


            sql = "SELECT dang_angle FROM wyb_unit_info WHERE dang_type = %s and type = %s"
            cursor.execute(sql, ( unit1, wyb_type ))
            db.commit()
            result = cursor.fetchone()

            begin_angle = float(result[0])



            sql = "SELECT unit FROM wyb_value WHERE id_num = %s and num = %s"
            cursor.execute(sql, ( i, wyb_num ))
            db.commit()
            result = cursor.fetchone()
            unit1 = str(result[0])



            print('处理前：' + 'unit1')

            unit1 = unit1.replace('+', '').replace('-', '')

            print('处理后：' + 'unit1')




            sql = "SELECT dang_angle FROM wyb_unit_info WHERE dang_type = %s and type = %s"
            cursor.execute(sql, (unit1, wyb_type))
            db.commit()
            result = cursor.fetchone()

            end_angle = float(result[0])


            print(begin_angle + '//' + end_angle)


            if (begin_angle == end_angle):

                pass

                print('不转')

            else:

                robot_act_new.execute_robot_rotate(begin_angle, end_angle, wyb_type)

                print('转')


            #  四、查询 万用表 测点等信息，通过函数 转换为 标准源 可读 测点

            sql = "SELECT unit,measure,hz FROM wyb_value WHERE id_num = %s and num = %s"

            cursor.execute(sql, (i, wyb_num))
            db.commit()
            result = cursor.fetchone()
            show_unit = str(result[0])
            mea_value = str(result[1])
            hz_value = int(result[2])

            mea_unit = self.run_check_num(show_unit)

            print(mea_value)





            #  五、 测量 该台 万用表

            if hz_value != 0:

                self.run_mea_same(show_unit, mea_unit, mea_value, vector_v1, bzy_v1, i, hz_value, wyb_num)

            else:

                self.run_mea_same(show_unit, mea_unit, mea_value, vector_v1, bzy_v1, i, 0, wyb_num)





            # 六、保存测量值，测量单位等信息，计算公式，随后导入数据库

            vector_v1[3] = show_unit
            vector_v1[4] = mea_value

            self.caculator_num(bzy_v1, vector_v1, i, wyb_num)

        '''


        #  七、转盘回至零位 OFF 档！！！

        sql = "SELECT dang_angle FROM wyb_unit_info WHERE dang_type = %s and type = %s"
        cursor.execute(sql, (unit1, wyb_type))
        db.commit()
        result = cursor.fetchone()
        begin_angle = float(result[0])

        unit2 = 'OFF'
        sql = "SELECT dang_angle FROM wyb_unit_info WHERE dang_type = %s and type = %s"
        cursor.execute(sql, (unit2, wyb_type))
        db.commit()
        result = cursor.fetchone()

        end_angle = float(result[0])

        robot_act_new.execute_robot_rotate(begin_angle, end_angle - 1, wyb_type)

        # 八、还原下一台万用表拨盘的初始位，使其上一台拨盘位归零，新万用表使用压力值校准拨盘位置

        sql = "UPDATE wyb_zp_temp SET zp_all_pos = %s WHERE id = 1"
        cursor.execute(sql, ('0&0&0&0&0&0',))
        db.commit()

        # 下一台！！！
        time.sleep(6)









    # 辅助函数

    def run_check_num(self, a):

        if 'M' in a and 'V' in a:
            return 'mV'
            # 检查字符串a是否含有'V'
        elif 'V' in a:
            return 'V'
            # 检查字符串a是否含有'A'且不含有'V'
        elif 'M' in a and 'A' in a:
            return 'mA'
            # 检查字符串a是否都含有'U'和'A'
        elif 'U' in a and 'A' in a:
            return 'uA'
            # 检查字符串a是否含有'R'和'K'
        elif 'A' in a and 'V' not in a:
            return 'A'
            # 检查字符串a是否都含有'M'和'A'
        elif 'R' in a and 'K' in a:
            return 'KOHM'
            # 检查字符串a是否含有'R'和'M'
        elif 'R' in a and 'M' in a:
            return 'MOHM'
            # 检查字符串a是否含有'R'且不含有'M'和'K'
        elif 'R' in a and 'M' not in a and 'K' not in a:
            return 'OHM'

    def run_find_dang_angle(self, begin, end, num):

        cursor = db.cursor()
        sql = "SELECT dang_angle FROM wyb_unit_info WHERE dang_num = %s and type = %s"

        cursor.execute(sql, (begin, num))
        db.commit()
        result = cursor.fetchone()
        begin_angle = float(result[0])

        cursor.execute(sql, (end, num))
        db.commit()
        result = cursor.fetchone()
        end_angle = float(result[0])

        return begin_angle, end_angle



    # 连续测量三次

    def run_mea_same(self, show_unit, mea_unit, mea_value, vector_v, bzy_v, table_lie, hz, wyb_num):

        self.ui.run_wyb_label_unit.setText(show_unit)
        self.ui.run_wyb_label_value.setText(str(mea_value))
        self.ui.run_wyb_label_result.setText(' ')
        # 测量值，测量单位，表格第几行，表格第几列，几位数保留小数位几位
        vector_v[0], bzy_v[0] = self.run_mea_unit(mea_value, mea_unit, 2, table_lie, hz)
        self.ui.run_wyb_label_result.setText(str(vector_v[0]))
        # 写入数据库
        cursor = db.cursor()
        sql1 = "UPDATE wyb_value SET m_1 = %s WHERE unit = %s AND measure = %s AND num = %s"
        cursor.execute(sql1, (str(vector_v[0]), show_unit, str(mea_value), wyb_num))
        db.commit()
        # 写入时间
        sql1 = "UPDATE wyb_value SET mea_time = %s WHERE unit = %s AND measure = %s AND num = %s"
        cursor.execute(sql1, ((datetime.now()).strftime("%Y%m%d%H%M%S"), show_unit, str(mea_value), wyb_num))
        db.commit()

        time.sleep(2)

        self.ui.run_wyb_label_unit.setText(show_unit)
        self.ui.run_wyb_label_value.setText(str(mea_value))
        self.ui.run_wyb_label_result.setText(' ')
        # 测量值，测量单位，表格第几行，表格第几列，几位数保留小数位几位
        vector_v[1], bzy_v[1] = self.run_mea_unit(mea_value, mea_unit, 3, table_lie, hz)
        self.ui.run_wyb_label_result.setText(str(vector_v[1]))
        # 写入数据库
        sql1 = "UPDATE wyb_value SET m_2 = %s WHERE unit = %s AND measure = %s AND num = %s"
        cursor.execute(sql1, (str(vector_v[1]), show_unit, str(mea_value), wyb_num))
        db.commit()

        time.sleep(2)

        self.ui.run_wyb_label_unit.setText(show_unit)
        self.ui.run_wyb_label_value.setText(str(mea_value))
        self.ui.run_wyb_label_result.setText(' ')
        # 测量值，测量单位，表格第几行，表格第几列，几位数保留小数位几位
        vector_v[2], bzy_v[2] = self.run_mea_unit(mea_value, mea_unit, 4, table_lie, hz)
        self.ui.run_wyb_label_result.setText(str(vector_v[2]))
        # 写入数据库
        sql1 = "UPDATE wyb_value SET m_3 = %s WHERE unit = %s AND measure = %s AND num = %s"
        cursor.execute(sql1, (str(vector_v[2]), show_unit, str(mea_value), wyb_num))
        db.commit()

        time.sleep(2)



    # 主 测量 函数
    '''
    def run_mea_unit_test(self, num, mea, hang, lie, hz):

        if num == '200' and mea == 'mA':

            if hz == 0:
                new_bzy.mea_dc('20mA')
                time.sleep(3)
                new_bzy.close_door()
                time.sleep(2)

        # 给标准源输出 30V 电压，更新表格
        if hz == 0:
            mea_value = str(num) + mea
            print('测量单位' + mea_value) 
            bzy_mea = new_bzy.mea_dc(mea_value)
        elif hz != 0:
            mea_value = str(num) + mea
            print('测量单位' + mea_value)
            print('hz单位' + hz)
            bzy_mea = new_bzy.mea_ac(mea_value, hz)

        bzy_mea = float(bzy_mea)
        time.sleep(4)

        # 识别 lcd 数字
        photo_name = str(num) + mea + '_' + datetime.now().strftime("%Y%m%d%H%M%S")
        testGrabImage.paizhao(photo_name)
        time.sleep(2)
        testGrabImage.paizhao(photo_name)
        time.sleep(1)
        self.run_image = "E:/photo/" + photo_name + ".jpg"

        mea_30 = ewm_and_lcd_formal.shuzishibie("E:/photo/" + photo_name + ".jpg")

        if ('L' not in mea_30) and ((mea_30 == 'lack_of_form') or (mea_30 == 'locate_incorrect') or (len(mea_30) != 4)):

            time.sleep(2)
            photo_name = str(num) + mea + '_' + datetime.now().strftime("%Y%m%d%H%M%S")
            testGrabImage.paizhao(photo_name)
            time.sleep(1)
            self.run_image = "E:/photo/" + photo_name + ".jpg"
            mea_30 = ewm_and_lcd_formal.shuzishibie("E:/photo/" + photo_name + ".jpg", -15)
            if ('L' not in mea_30) and (
                    (mea_30 == 'lack_of_form') or (mea_30 == 'locate_incorrect') or (len(mea_30) != 4)):
                mea_30 = ewm_and_lcd_formal.shuzishibie("E:/photo/" + photo_name + ".jpg", -16)
                if ('L' not in mea_30) and (
                        (mea_30 == 'lack_of_form') or (mea_30 == 'locate_incorrect') or (len(mea_30) != 4)):
                    mea_30 = ewm_and_lcd_formal.shuzishibie("E:/photo/" + photo_name + ".jpg", -18)
                    if ('L' not in mea_30) and (
                            (mea_30 == 'lack_of_form') or (mea_30 == 'locate_incorrect') or (len(mea_30) != 4)):
                        mea_30 = ewm_and_lcd_formal.shuzishibie("E:/photo/" + photo_name + ".jpg", -19)
                        if ('L' not in mea_30) and (
                                (mea_30 == 'lack_of_form') or (mea_30 == 'locate_incorrect') or (len(mea_30) != 4)):
                            mea_30 = ewm_and_lcd_formal.shuzishibie("E:/photo/" + photo_name + ".jpg", -17, 1)

        print('OL有吗')
        print(mea_30)

        delta_bzy_mea = 0

        if ('L' in mea_30) :

            print('超！')
            new_bzy.close_door()

            newItem = QTableWidgetItem('超量程')
            self.ui.run_table.setItem(12, lie, newItem)

            if '-' in num:

                len_value = len(num) - 1
                num = int(num)

                if len_value == 3:
                    num = num + 1
                    delta_bzy_mea = -1
                elif len_value == 2:
                    num = num + 0.1
                    delta_bzy_mea = -0.1
                elif len_value == 1:
                    num = num + 0.01
                    delta_bzy_mea = -0.01
                elif len_value == 4:
                    num = num + 10
                    delta_bzy_mea = -10

            else:

                len_value = len(num)
                num = int(num)

                if len_value == 3:
                    num = num - 1
                    delta_bzy_mea = 1
                elif len_value == 2:
                    num = num - 0.1
                    delta_bzy_mea = 0.1
                elif len_value == 1:
                    num = num - 0.01
                    delta_bzy_mea = 0.01
                elif len_value == 4:
                    num = num - 10
                    delta_bzy_mea = 10

            time.sleep(4)

            if hz == 0:

                mea_value = str(num) + mea
                print('测量单位')
                print(mea_value)
                bzy_mea = new_bzy.mea_dc(mea_value)

            elif hz != 0:

                mea_value = str(num) + mea
                print('hz单位')
                print(mea_value)
                bzy_mea = new_bzy.mea_ac(mea_value, hz)

            bzy_mea = float(bzy_mea)
            time.sleep(4)

            # 识别 lcd 数字
            photo_name = str(num) + mea + '_' + datetime.now().strftime("%Y%m%d%H%M%S")
            testGrabImage.paizhao(photo_name)
            time.sleep(1)
            self.run_image = "E:/photo/" + photo_name + ".jpg"

            mea_30 = ewm_and_lcd_formal.shuzishibie("E:/photo/" + photo_name + ".jpg")

            if ('L' not in mea_30) and (
                    (mea_30 == 'lack_of_form') or (mea_30 == 'locate_incorrect') or (len(mea_30) != 4)):
                time.sleep(2)
                photo_name = str(num) + mea + '_' + datetime.now().strftime("%Y%m%d%H%M%S")
                testGrabImage.paizhao(photo_name)
                time.sleep(1)
                self.run_image = "E:/photo/" + photo_name + ".jpg"
                mea_30 = ewm_and_lcd_formal.shuzishibie("E:/photo/" + photo_name + ".jpg", -15)
                if ('L' not in mea_30) and (
                        (mea_30 == 'lack_of_form') or (mea_30 == 'locate_incorrect') or (len(mea_30) != 4)):
                    mea_30 = ewm_and_lcd_formal.shuzishibie("E:/photo/" + photo_name + ".jpg", -16)
                    if ('L' not in mea_30) and (
                            (mea_30 == 'lack_of_form') or (mea_30 == 'locate_incorrect') or (len(mea_30) != 4)):
                        mea_30 = ewm_and_lcd_formal.shuzishibie("E:/photo/" + photo_name + ".jpg", -18)
                        if ('L' not in mea_30) and (
                                (mea_30 == 'lack_of_form') or (mea_30 == 'locate_incorrect') or (len(mea_30) != 4)):
                            mea_30 = ewm_and_lcd_formal.shuzishibie("E:/photo/" + photo_name + ".jpg", -19)
                            if ('L' not in mea_30) and (
                                    (mea_30 == 'lack_of_form') or (mea_30 == 'locate_incorrect') or (len(mea_30) != 4)):
                                mea_30 = ewm_and_lcd_formal.shuzishibie("E:/photo/" + photo_name + ".jpg", -17, 1)

        if mea_30 == 'locate_incorrect':
            newItem = QTableWidgetItem('定位失败')
            self.ui.run_table.setItem(11, lie, newItem)
            mea_30 = 1

        print('最初始LCD上的4位mea_30')
        print(mea_30)

        if 'L' in mea_30:
            newItem = QTableWidgetItem('不合格')
            self.ui.run_table.setItem(11, lie, newItem)
            mea_30 = -1

        if mea_30 == '0000':
            newItem = QTableWidgetItem('该档位烧掉')
            self.ui.run_table.setItem(11, lie, newItem)
            self.ui.run_table.setItem(12, lie, newItem)
            mea_30 = -1

        # 处理 lcd 读数，加小数点，并显示于 ui 界面上
        mea_30 = int(mea_30)
        mea_30 = float(mea_30)

        xunhuan = 0

        while (True):

            print('bzy_mea')
            print(bzy_mea)
            print('mea_30')
            print(mea_30)
            value_test = mea_30 / bzy_mea

            if (value_test > 0.8 and value_test < 1.2) or (value_test < -0.8 and value_test > -1.2):

                break

            else:

                mea_30 = mea_30 * 0.1

                xunhuan = xunhuan + 1

                if xunhuan > 5:
                    print('sbsb')
                    mea_30 = mea_30 * 100000
                    break

        str_mea_num = len(str(int(bzy_mea)))

        if '-' in str(int(bzy_mea)):
            str_mea_num = str_mea_num - 1
            mea_30 = - mea_30

        print('补偿后的mea30')
        print(mea_30)

        if delta_bzy_mea == -0.1:
            mea_30 = mea_30 - 0.1
            bzy_mea = bzy_mea - 0.1
        elif delta_bzy_mea == -1:
            mea_30 = mea_30 - 1
            bzy_mea = bzy_mea - 1
        elif delta_bzy_mea == -0.01:
            mea_30 = mea_30 - 0.01
            bzy_mea = bzy_mea - 0.01
        elif delta_bzy_mea == -10:
            mea_30 = mea_30 - 10
            bzy_mea = bzy_mea - 10
        elif delta_bzy_mea == 0.1:
            mea_30 = mea_30 + 0.1
            bzy_mea = bzy_mea + 0.1
        elif delta_bzy_mea == 1:
            mea_30 = mea_30 + 1
            bzy_mea = bzy_mea + 1
        elif delta_bzy_mea == 10:
            mea_30 = mea_30 + 10
            bzy_mea = bzy_mea + 10
        elif delta_bzy_mea == 0.01:
            mea_30 = mea_30 + 0.01
            bzy_mea = bzy_mea + 0.01

        print('还原后的mea30')
        print(mea_30)

        # print('str_mea_num')
        # print(str_mea_num)

        mea_30 = format(mea_30, f'.{(4 - str_mea_num)}f')

        print('去掉多余位数的mea30')
        print(mea_30)

       

        newItem = QTableWidgetItem(str(mea_30))
        self.ui.run_table.setItem(hang, lie, newItem)

        # 切换lcd图片
        self.alter_image()

        time.sleep(1)
        # 关闭标准源通道
        new_bzy.close_door()

        time.sleep(2)

        return mea_30, bzy_mea
    '''

    def run_mea_unit(self, num, mea, hang, lie, hz):


        time.sleep(2)

        # 给标准源输出 30V 电压，更新表格
        if hz == 0:
            mea_value = str(num) + mea
            print('测量单位' + str(mea_value))
            bzy_mea = new_bzy.mea_dc(mea_value)
        elif hz != 0:
            mea_value = str(num) + mea
            print('测量单位' + str(mea_value))
            print('hz单位' + str(hz))
            bzy_mea = new_bzy.mea_ac(mea_value, hz)

        bzy_mea = float(bzy_mea)
        time.sleep(4)

        # 识别 lcd 数字
        photo_name = str(num) + mea + '_' + datetime.now().strftime("%Y%m%d%H%M%S")
        testGrabImage.paizhao(photo_name, 2, photo_num)
        time.sleep(2)
        testGrabImage.paizhao(photo_name, 2, photo_num)
        time.sleep(1)
        self.run_image = "E:/photo/p" + str(photo_num) + "/" + photo_name + ".jpg"

        mea_30 = ewm_and_lcd_formal.shuzishibie("E:/photo/p" + str(photo_num) + "/" + photo_name + ".jpg")

        print('LCD上的原值：' + str(mea_30))

        if ('L' not in mea_30) and ((len(mea_30) != 4)):

            if len(mea_30) == 3:
                mea_30 = '1' + mea_30
            elif len(mea_30) == 2:
                mea_30 = '11' + mea_30
            elif len(mea_30) == 1:
                mea_30 = '111' + mea_30

            print('补1后的值：' + str(mea_30))

        delta_bzy_mea = 0

        # if ('L' in mea_30) or (num == '200' and mea == 'mA' and hz == 0) or (num == '-200' and mea == 'mA' and hz == 0):

        if 'L' in mea_30:

            print('进入超量程补偿！')
            new_bzy.close_door()

            newItem = QTableWidgetItem('超量程')
            self.ui.run_table.setItem(12, lie, newItem)

            if '-' in num:

                len_value = len(num) - 1
                num = int(num)

                if len_value == 3:
                    num = num + 1
                    delta_bzy_mea = -1
                elif len_value == 2:
                    num = num + 0.1
                    delta_bzy_mea = -0.1
                elif len_value == 1:
                    num = num + 0.01
                    delta_bzy_mea = -0.01
                elif len_value == 4:
                    num = num + 10
                    delta_bzy_mea = -10

            else:

                len_value = len(num)
                num = int(num)

                if len_value == 3:
                    num = num - 1
                    delta_bzy_mea = 1
                elif len_value == 2:
                    num = num - 0.1
                    delta_bzy_mea = 0.1
                elif len_value == 1:
                    num = num - 0.01
                    delta_bzy_mea = 0.01
                elif len_value == 4:
                    num = num - 10
                    delta_bzy_mea = 10

            time.sleep(2)

            if hz == 0:

                mea_value = str(num) + mea
                bzy_mea = new_bzy.mea_dc(mea_value)

            elif hz != 0:

                mea_value = str(num) + mea
                bzy_mea = new_bzy.mea_ac(mea_value, hz)

            bzy_mea = float(bzy_mea)
            time.sleep(4)

            # 识别 lcd 数字
            photo_name = str(num) + mea + '_' + datetime.now().strftime("%Y%m%d%H%M%S")
            testGrabImage.paizhao(photo_name, 2, photo_num)
            time.sleep(1)
            self.run_image = "E:/photo/p" + str(photo_num) + "/" + photo_name + ".jpg"

            mea_30 = ewm_and_lcd_formal.shuzishibie("E:/photo/p" + str(photo_num) + "/" + photo_name + ".jpg")

            print('第二次LCD上的原值：' + str(mea_30))

            if ('L' not in mea_30) and ((len(mea_30) != 4)):

                if len(mea_30) == 3:
                    mea_30 = '1' + mea_30
                elif len(mea_30) == 2:
                    mea_30 = '11' + mea_30
                elif len(mea_30) == 1:
                    mea_30 = '111' + mea_30

                print('补1后的值：' + str(mea_30))

        not_buchang = 0

        if 'L' in mea_30:

            newItem1 = QTableWidgetItem('不合格')
            self.ui.run_table.setItem(11, lie, newItem1)

            newItem2 = QTableWidgetItem('两次超量程')
            self.ui.run_table.setItem(12, lie, newItem2)

            if int(num) < 0:
                buchang = -1
            else:
                buchang = 1

            mea_30 = int(num) + buchang * 0.31 * int(num)

            not_buchang = 1


        if mea_30 == '0000':
            newItem = QTableWidgetItem('该档位烧掉')
            self.ui.run_table.setItem(11, lie, newItem)
            self.ui.run_table.setItem(12, lie, newItem)
            mea_30 = -1
            not_buchang = 1


        if not_buchang == 0:

            # 处理 lcd 读数，加小数点，并显示于 ui 界面上
            mea_30 = int(mea_30)
            mea_30 = float(mea_30)

            xunhuan = 0

            while (True):

                #print('bzy_mea')
                #print(bzy_mea)
                #print('mea_30')
                #print(mea_30)

                value_test = mea_30 / bzy_mea

                if (value_test > 0.8 and value_test < 1.2) or (value_test < -0.8 and value_test > -1.2):

                    break

                else:

                    mea_30 = mea_30 * 0.1

                    xunhuan = xunhuan + 1

                    if xunhuan > 5:
                        print('数据异常，补偿值×100000')
                        mea_30 = mea_30 * 100000
                        break

            str_mea_num = len(str(int(bzy_mea)))

            if '-' in str(int(bzy_mea)):
                str_mea_num = str_mea_num - 1
                mea_30 = - mea_30

            print('加上小数位的LCD值：' + str(mea_30))


            if delta_bzy_mea == -0.1:
                mea_30 = mea_30 - 0.1
                bzy_mea = bzy_mea - 0.1
            elif delta_bzy_mea == -1:
                mea_30 = mea_30 - 1
                bzy_mea = bzy_mea - 1
            elif delta_bzy_mea == -0.01:
                mea_30 = mea_30 - 0.01
                bzy_mea = bzy_mea - 0.01
            elif delta_bzy_mea == -10:
                mea_30 = mea_30 - 10
                bzy_mea = bzy_mea - 10
            elif delta_bzy_mea == 0.1:
                mea_30 = mea_30 + 0.1
                bzy_mea = bzy_mea + 0.1
            elif delta_bzy_mea == 1:
                mea_30 = mea_30 + 1
                bzy_mea = bzy_mea + 1
            elif delta_bzy_mea == 10:
                mea_30 = mea_30 + 10
                bzy_mea = bzy_mea + 10
            elif delta_bzy_mea == 0.01:
                mea_30 = mea_30 + 0.01
                bzy_mea = bzy_mea + 0.01


            mea_30 = format(mea_30, f'.{(4 - str_mea_num)}f')

            wucha = (float(mea_30) - float(num))/float(num) * 100

            if wucha <= -15.0 or wucha >= 15.0:

                source_file_path = "E:/photo/p" + str(photo_num) + "/" + photo_name + ".jpg"

                destination_folder_path = "E:/photo/problem"

                # 构建目标文件路径
                destination_file_path = os.path.join(destination_folder_path, f"{photo_name}.jpg")

                # 复制文件
                shutil.copy(source_file_path, destination_file_path)





        elif not_buchang == 1:

            str_mea_num = len(str(int(bzy_mea)))

            if '-' in str(int(bzy_mea)):
                str_mea_num = str_mea_num - 1

            mea_30 = format(mea_30, f'.{(4 - str_mea_num)}f')



        newItem = QTableWidgetItem(str(mea_30))
        self.ui.run_table.setItem(hang, lie, newItem)

        # 切换lcd图片
        self.alter_image()

        time.sleep(1)
        # 关闭标准源通道
        new_bzy.close_door()



        return mea_30, bzy_mea



    # 计算

    def caculator_num(self, vec_real, vec_mea, lie, wyb_num):

        cursor = db.cursor()

        # 一、 通过 辅助函数 得到 万用表 该测点 测量值 的 小数位数

        point_num = int(self.count_decimal_places(vec_mea[0]))

        vec_mea[0] = float(vec_mea[0])
        vec_mea[1] = float(vec_mea[1])
        vec_mea[2] = float(vec_mea[2])

        # 二、 计算 平均值，最大，最小，示值误差

        self.ave_mea = round(((vec_mea[0] + vec_mea[1] + vec_mea[2]) / 3), point_num)
        self.max_mea = max([vec_mea[0], vec_mea[1], vec_mea[2]])
        self.min_mea = min([vec_mea[0], vec_mea[1], vec_mea[2]])
        self.vec_real_num = round(((vec_real[0] + vec_real[1] + vec_real[2]) / 3), 4)
        self.minus_mea = self.ave_mea - self.vec_real_num

        # 三、 当 测点 读数 显示 异常时， 将 示值误差 和 相对误差 显示为 100% ， 正常时 小数位保留 三位

        if self.ave_mea == 0 or self.vec_real_num == 0:

            self.rel_mea = 100.000
            self.rep_mea = 100.000

        else:

            self.rel_mea = float(format(((self.ave_mea - self.vec_real_num) / abs(self.vec_real_num)) * 100, f'.3f'))
            self.rep_mea = float(format(((self.max_mea - self.min_mea) / abs(self.ave_mea)) * 100, f'.3f'))

        #  四、 当 是否超量程 显示 超量程 或 该档位烧掉时 选择：
        # 超量程：数据库ol字段更新，根据电流电压电阻，计算是否 合格，显示于 表中的 结果
        # 其他情况： 更新 数据库 写入参数
        # 不合格情况：——结果：不合格；——是否超量程：超量程，那么更新数据库ol为超量程，并且通过 不合格 给定测点 数值，计算大概率为 不合格，显示于表中
        # 当 正常 时：数据库ol字段更新，根据表中 测点，返回 电流电压电阻 标志 参数，计算是否 合格

        if self.ui.run_table.item(12, lie) is not None:

            if self.ui.run_table.item(12, lie).text() == '超量程':

                sql_result = 'overload'
                sql1 = "UPDATE wyb_value SET ol = %s WHERE unit = %s AND measure = %s AND num = %s"
                cursor.execute(sql1, (str(sql_result), vec_mea[3], str(vec_mea[4]), wyb_num))
                db.commit()

                taitou = self.run_check_num(str(self.ui.run_table.item(0, lie).text()))

                if 'V' in taitou:
                    zhi = 1.0
                elif ('OHM' in taitou) or ('A' in taitou):
                    zhi = 0.5

                if self.rel_mea >= (-1.0 * zhi) and self.rel_mea <= (1.0 * zhi):
                    mea_result = '合格'
                    sql_result = 'pass'
                    self.ui.run_table.setItem(11, lie, QTableWidgetItem(mea_result))
                else:
                    mea_result = '不合格'
                    sql_result = 'nopass'
                    self.ui.run_table.setItem(11, lie, QTableWidgetItem(mea_result))

            elif (self.ui.run_table.item(11, lie).text() == '不合格') or \
                    (self.ui.run_table.item(12, lie).text() == '两次超量程'):
                sql_result = 'nopass'
            elif self.ui.run_table.item(11, lie).text() == '定位失败':
                sql_result = 'lcd_incorrect'
            elif self.ui.run_table.item(11, lie).text() == '该档位烧掉':
                sql_result = 'bad'

        else:

            self.ui.run_table.setItem(12, lie, QTableWidgetItem('正常'))
            sql1 = "UPDATE wyb_value SET ol = %s WHERE unit = %s AND measure = %s AND num = %s"
            cursor.execute(sql1, (str('ok'), vec_mea[3], str(vec_mea[4]), wyb_num))
            db.commit()

            taitou = self.run_check_num(str(self.ui.run_table.item(0, lie).text()))

            if 'V' in taitou:
                zhi = 1.0
            elif ('OHM' in taitou) or ('A' in taitou):
                zhi = 0.5

            if self.rel_mea >= (-1.0 * zhi) and self.rel_mea <= (1.0 * zhi):
                mea_result = '合格'
                sql_result = 'pass'
                self.ui.run_table.setItem(11, lie, QTableWidgetItem(mea_result))
            else:
                mea_result = '不合格'
                sql_result = 'nopass'
                self.ui.run_table.setItem(11, lie, QTableWidgetItem(mea_result))

        # 五、保留 计算结果的 小数位 数

        self.ave_mea = format(self.ave_mea, f'.{(point_num)}f')
        self.max_mea = format(self.max_mea, f'.{(point_num)}f')
        self.min_mea = format(self.min_mea, f'.{(point_num)}f')
        self.minus_mea = format(self.minus_mea, f'.{(point_num)}f')

        self.rel_mea = format(self.rel_mea, f'.3f') + '%'
        self.rep_mea = format(self.rep_mea, f'.3f') + '%'

        # 六、 更新 数据库 计算值，结果 值，随后更新表格 的 计算值

        sql1 = "UPDATE wyb_value SET m_average = %s WHERE unit = %s AND measure = %s AND num = %s"
        cursor.execute(sql1, (str(self.ave_mea), vec_mea[3], str(vec_mea[4]), wyb_num))
        db.commit()
        sql1 = "UPDATE wyb_value SET m_max = %s WHERE unit = %s AND measure = %s AND num = %s"
        cursor.execute(sql1, (str(self.max_mea), vec_mea[3], str(vec_mea[4]), wyb_num))
        db.commit()
        sql1 = "UPDATE wyb_value SET m_max = %s WHERE unit = %s AND measure = %s AND num = %s"
        cursor.execute(sql1, (str(self.max_mea), vec_mea[3], str(vec_mea[4]), wyb_num))
        db.commit()
        sql1 = "UPDATE wyb_value SET m_min = %s WHERE unit = %s AND measure = %s AND num = %s"
        cursor.execute(sql1, (str(self.min_mea), vec_mea[3], str(vec_mea[4]), wyb_num))
        db.commit()
        sql1 = "UPDATE wyb_value SET m_sz = %s WHERE unit = %s AND measure = %s AND num = %s"
        cursor.execute(sql1, (str(self.minus_mea), vec_mea[3], str(vec_mea[4]), wyb_num))
        db.commit()
        sql1 = "UPDATE wyb_value SET m_xd = %s WHERE unit = %s AND measure = %s AND num = %s"
        cursor.execute(sql1, (str(self.rel_mea), vec_mea[3], str(vec_mea[4]), wyb_num))
        db.commit()
        sql1 = "UPDATE wyb_value SET m_repeat = %s WHERE unit = %s AND measure = %s AND num = %s"
        cursor.execute(sql1, (str(self.rep_mea), vec_mea[3], str(vec_mea[4]), wyb_num))
        db.commit()

        sql1 = "UPDATE wyb_value SET pass = %s WHERE unit = %s AND measure = %s AND num = %s"
        cursor.execute(sql1, (str(sql_result), vec_mea[3], str(vec_mea[4]), wyb_num))
        db.commit()

        self.ui.run_table.setItem(5, lie, QTableWidgetItem(str(self.ave_mea)))
        self.ui.run_table.setItem(6, lie, QTableWidgetItem(str(self.max_mea)))
        self.ui.run_table.setItem(7, lie, QTableWidgetItem(str(self.min_mea)))
        self.ui.run_table.setItem(8, lie, QTableWidgetItem(str(self.minus_mea)))
        self.ui.run_table.setItem(9, lie, QTableWidgetItem(self.rel_mea))
        self.ui.run_table.setItem(10, lie, QTableWidgetItem(self.rep_mea))

    # 其他

    def alter_image(self):

        self.pixmap = QPixmap(self.run_image)
        self.new_pixmap = self.pixmap.scaled(self.ui.run_graph.width(), self.ui.run_graph.height())
        self.ui.run_graph.setPixmap(self.new_pixmap)

    def alter_kong_image(self, manual_image):

        self.pixmap = QPixmap(manual_image)
        self.new_pixmap = self.pixmap.scaled(self.ui.run_manual_graph.width(), self.ui.run_manual_graph.height())
        self.ui.run_manual_graph.setPixmap(self.new_pixmap)

    def run_clear_wyb(self):

        for i in range(39):

            for j in range(13):
                self.run_table.setItem(j, i + 1, None)

        self.ui.run_wyb_label_unit.setText(' ')
        self.ui.run_wyb_label_value.setText(' ')
        self.ui.run_wyb_label_result.setText(' ')

    def run_exit(self):

        self.zhujiemian = wyb()
        self.zhujiemian.ui.show()
        self.ui.close()

    def run_stop(self):

        pid = os.getpid()  # 获取当前进程的PID
        os.kill(pid, signal.SIGTERM)  # 发
        robot_act_new.sys.exit("1")

    def count_decimal_places(self, num):
        # 将数字转换为字符串

        str_num = str(num)
        # 找到小数点的位置
        decimal_point_index = str_num.find('.')
        # 如果没有小数点，返回0
        if decimal_point_index == -1:
            return 0
        # 计算小数点后的字符数量
        decimal_places = len(str_num) - decimal_point_index - 1
        return decimal_places



    # 单测 200 ma 档 连续 三次
    def run_mea_vc_test(self, wyb_num):

        cursor = db.cursor()

        sql = "SELECT unit,measure,hz FROM wyb_value WHERE id_num = %s and num = %s"

        cursor.execute(sql, (21, 'VICTORVC890C+001'))
        db.commit()
        result = cursor.fetchone()
        show_unit = str(result[0])
        mea_value = str(result[1])
        hz_value = int(result[2])

        mea_unit = self.run_check_num(show_unit)

        print(mea_value)
        print(1)
        print(mea_unit)
        if hz_value != 0:

            self.run_mea_same(show_unit, mea_unit, '200', vector_v1, bzy_v1, 21, hz_value)

        else:

            self.run_mea_same(show_unit, mea_unit, '200', vector_v1, bzy_v1, 21, 0, )

        # 记录测量单位
        vector_v1[3] = show_unit
        vector_v1[4] = mea_value

        self.caculator_num(bzy_v1, vector_v1, 21, wyb_num)
